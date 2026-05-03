# Standard library imports
import os
import base64
import json
import time
import hashlib
import secrets
from io import BytesIO
from datetime import timedelta

# Django imports
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, EmailMessage
from django.db.models import Q
from django.forms import ValidationError
from django.http import Http404, HttpResponse, HttpResponseNotFound, JsonResponse, StreamingHttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.views.decorators.http import require_POST
from django.http import HttpResponse

# Third party imports
from loguru import logger
import razorpay
from razorpay.errors import SignatureVerificationError
# from minio import S3Error  # Commented out - no longer using MinIO

# Local application imports
from .forms import UserLoginForm, UserRegistrationForm, UserUpdateForm
from .models import Announcement, OTP, User, CategoryChangeRequest, RegistrationQuestion, RegistrationAnswer, SystemSettings, PaymentRegistration
from .utils import (
    delete_file_from_minio,
    fetch_file_from_minio,
    generate_qr_code_live,
    generate_random_password,
    get_gallery_images,
    member_profile_file_gaps,
    MEMBER_PROFILE_FILE_FIELDS,
    send_email_with_attachments,
    send_new_account_credentials_email,
    send_otp_via_email,
    stored_file_is_missing,
    upload_file_to_minio,
)
from .bulk_user_import import build_bulk_user_template_bytes, import_users_from_xlsx
from .id_card_image import build_id_card_png_bytes, id_card_png_filename
from app.decorators import admin_required

# Prefer env so Dashboard key + secret always match (signature verify uses secret).
RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "rzp_live_SeUfSkspvWpRZe")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", "dIzu8wg41csqXnm7eBFy7Eik")

MEMBERSHIP_TYPES = {
    "adhoc": {
        "label": "Adhoc",
        "amount": 200000,  # ₹2,000 (paise for Razorpay)
    },
    "lifetime": {
        "label": "Lifetime",
        "amount": 250000,  # ₹2,500 (paise for Razorpay)
    },
    "amc": {
        "label": "AMC",
        "amount": 50000,  # ₹500 (paise for Razorpay)
    },
}


def _is_amc_payment_registration(payment) -> bool:
    """True for seeded AMC rows (receipt prefix) or notes flag."""
    if not payment:
        return False
    r = (getattr(payment, "receipt", None) or "").strip().upper()
    if r.startswith("AMC-"):
        return True
    notes = getattr(payment, "notes", None) or {}
    if isinstance(notes, dict) and (notes.get("membership_type") or "").strip().lower() == "amc":
        return True
    return False


def _user_category_choice_tuples():
    """All valid User.category (value, label) pairs — legacy + adhoc + lifetime."""
    return list(User._meta.get_field("category").choices)


def _user_category_values():
    return {c[0] for c in _user_category_choice_tuples()}


def _portal_session_user(request):
    """
    Resolve the logged-in member the same way the portal home page does: session
    username first (set at login), then Django authenticated user.
    """
    uname = request.session.get("username")
    if uname:
        u = User.objects.filter(username=uname).first()
        if u:
            return u
    if request.user.is_authenticated:
        return User.objects.filter(pk=request.user.pk).first()
    return None


def _payment_registration_visible_to_user(payment, user):
    """True if this payment row belongs to the portal user (FK or same email for legacy rows)."""
    if not user or not payment:
        return False
    if payment.user_id:
        return int(payment.user_id) == int(user.pk)
    uemail = (user.email or "").strip().lower()
    pemail = (payment.email or "").strip().lower()
    return bool(uemail and pemail and uemail == pemail)


def _payment_registration_belongs_to_member(payment, member):
    """True if this payment row is for the given User (FK or same email for legacy rows)."""
    if not member or not payment:
        return False
    if payment.user_id:
        return int(payment.user_id) == int(member.pk)
    memail = (member.email or "").strip().lower()
    pemail = (payment.email or "").strip().lower()
    return bool(memail and pemail and memail == pemail)


def _payment_receipt_kind_label(payment):
    notes = payment.notes if isinstance(payment.notes, dict) else {}
    if (notes.get("membership_type") or "").strip().lower() == "amc" or (
        (payment.receipt or "").strip().upper().startswith("AMC-")
    ):
        return "AMC (annual)"
    return (notes.get("membership_label") or "").strip() or "Membership / fee"


def _category_display_name(value):
    if value is None or value == "":
        return ""
    return dict(_user_category_choice_tuples()).get(value, value)


def _get_membership_details(membership_type):
    details = MEMBERSHIP_TYPES.get((membership_type or "").strip().lower())
    if not details:
        return None

    return {
        **details,
        "amount_rupees": details["amount"] // 100,
    }


def _get_razorpay_client():
    return razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))


RZP_PENDING_SESSION_TTL_SEC = 15 * 60


def _rzp_checkout_session_clear(request):
    request.session.pop("rzp_pending_registration_id", None)
    request.session.pop("rzp_pending_expires", None)


def _rzp_checkout_session_set(request, registration_pk):
    """Remember pending checkout so refresh does not create duplicate Razorpay orders."""
    pk = int(registration_pk)
    prev = request.session.get("rzp_pending_registration_id")
    if prev is not None and int(prev) != pk:
        PaymentRegistration.objects.filter(pk=int(prev), payment_status="pending").update(
            payment_status="failed"
        )
    request.session["rzp_pending_registration_id"] = pk
    request.session["rzp_pending_expires"] = timezone.now().timestamp() + RZP_PENDING_SESSION_TTL_SEC


def health(request):
    return HttpResponse("OK")
	
def get_registration(request):
    # AMC is paid from the member profile / pay-amc flow, not the public signup form.
    membership_types = [
        (value, details["label"], details["amount"] // 100)
        for value, details in MEMBERSHIP_TYPES.items()
        if value != "amc"
    ]
    return render(
        request,
        "registration.html",
        {
            "membership_types": membership_types,
        },
    )


@require_POST
def create_order(request):
    name = request.POST.get("name", "").strip()
    email = request.POST.get("email", "").strip()
    phone = request.POST.get("phone", "").strip()
    membership_type = request.POST.get("membership_type", "").strip().lower()
    membership_details = _get_membership_details(membership_type)

    if not all([name, email, phone]):
        return JsonResponse({"error": "Name, email and phone are required."}, status=400)
    if not membership_details:
        return JsonResponse({"error": "A valid membership type is required."}, status=400)

    # Razorpay amounts must be integer paise (INR).
    amount = int(membership_details["amount"])

    linked_user = User.objects.filter(email=email).first()
    # Placeholder receipt: model requires unique non-empty receipt before Razorpay receipt is assigned.
    placeholder_receipt = f"pnd_{secrets.token_hex(12)}"[:40]
    registration = PaymentRegistration.objects.create(
        user=linked_user,
        name=name,
        email=email,
        phone=phone,
        amount=amount,
        currency="INR",
        receipt=placeholder_receipt,
        payment_status="pending",
        notes={
            "membership_type": membership_type,
            "membership_label": membership_details["label"],
        },
    )

    # Receipt sent to Razorpay: unique, max 40 chars (API limit).
    receipt = f"r{registration.pk}_{secrets.token_hex(3)}"[:40]

    # Notes values must be strings for Razorpay.
    rz_notes = {
        "registration_id": str(registration.id),
        "membership_type": str(membership_type),
        "membership_label": str(membership_details["label"]),
    }

    data = {
        "amount": amount,
        "currency": "INR",
        "receipt": receipt,
        "notes": rz_notes,
    }

    try:
        client = _get_razorpay_client()
        order = client.order.create(data=data)
    except Exception as exc:
        logger.error(f"Razorpay order creation failed: {exc}")
        registration.payment_status = "failed"
        registration.receipt = receipt
        registration.save(update_fields=["payment_status", "receipt", "updated_at"])
        return JsonResponse({"error": "Unable to create payment order."}, status=500)

    order_id = str(order.get("id") or "").strip()
    if not order_id:
        logger.error("Razorpay order.create returned no id: %s", order)
        registration.payment_status = "failed"
        registration.receipt = receipt
        registration.save(update_fields=["payment_status", "receipt", "updated_at"])
        return JsonResponse({"error": "Payment gateway returned an invalid order."}, status=500)

    registration.razorpay_order_id = order_id
    registration.receipt = receipt
    registration.save(update_fields=["razorpay_order_id", "receipt", "updated_at"])

    logger.info(
        "Razorpay order created registration_id=%s order_id=%s amount_paise=%s receipt=%s",
        registration.pk,
        order_id,
        amount,
        receipt,
    )

    _rzp_checkout_session_set(request, registration.pk)

    return JsonResponse(
        {
            "order_id": order_id,
            "amount": int(order.get("amount") or amount),
            "currency": order.get("currency") or "INR",
            "key": RAZORPAY_KEY_ID,
            "registration_id": registration.id,
        }
    )


@require_POST
def verify_payment(request):
    registration_id = (request.POST.get("registration_id") or "").strip()
    razorpay_payment_id = (request.POST.get("razorpay_payment_id") or "").strip()
    razorpay_order_id = (request.POST.get("razorpay_order_id") or "").strip()
    razorpay_signature = (request.POST.get("razorpay_signature") or "").strip()

    logger.warning(
        "DEBUG PAYMENT raw POST registration_id=%r order_id=%r payment_id=%r signature_len=%s",
        registration_id,
        razorpay_order_id,
        razorpay_payment_id,
        len(razorpay_signature),
    )

    if not all([registration_id, razorpay_payment_id, razorpay_order_id, razorpay_signature]):
        return JsonResponse({"success": False, "error": "Missing payment verification fields."}, status=400)

    try:
        registration = PaymentRegistration.objects.get(pk=registration_id)
    except (PaymentRegistration.DoesNotExist, ValueError, TypeError):
        return JsonResponse({"success": False, "error": "Invalid or unknown registration."}, status=400)

    logger.warning(
        "DEBUG PAYMENT registration pk=%s status=%s stored_order=%r received_order=%r payment_id=%r",
        registration.pk,
        registration.payment_status,
        registration.razorpay_order_id,
        razorpay_order_id,
        razorpay_payment_id,
    )

    # Idempotent success if client retries verify after payment already captured.
    stored_pid = (registration.razorpay_payment_id or "").strip()
    if registration.payment_status == "paid" and stored_pid == razorpay_payment_id:
        logger.info("Razorpay verify idempotent OK registration_id=%s", registration_id)
        _rzp_checkout_session_clear(request)
        return JsonResponse({"success": True, "redirect_url": reverse("payment_success")})
    if registration.payment_status == "paid":
        _rzp_checkout_session_clear(request)
        return JsonResponse(
            {"success": False, "error": "This registration is already marked paid with a different payment id."},
            status=400,
        )

    stored_oid = (registration.razorpay_order_id or "").strip()
    if not stored_oid:
        logger.error("DEBUG PAYMENT missing stored order on registration_id=%s", registration_id)
        _rzp_checkout_session_clear(request)
        return JsonResponse({"success": False, "error": "Missing stored order for this registration."}, status=400)
    if stored_oid != razorpay_order_id:
        logger.warning(
            "Razorpay order mismatch: registration=%s stored=%r posted=%r",
            registration_id,
            registration.razorpay_order_id,
            razorpay_order_id,
        )
        registration.payment_status = "failed"
        registration.save(update_fields=["payment_status", "updated_at"])
        _rzp_checkout_session_clear(request)
        return JsonResponse(
            {
                "success": False,
                "error": "Order mismatch: this payment does not match the order created for your checkout session.",
            },
            status=400,
        )

    params_dict = {
        "razorpay_order_id": razorpay_order_id,
        "razorpay_payment_id": razorpay_payment_id,
        "razorpay_signature": razorpay_signature,
    }
    client = _get_razorpay_client()
    try:
        client.utility.verify_payment_signature(params_dict)
    except SignatureVerificationError as exc:
        logger.warning(
            "Razorpay signature verification failed (registration=%s order=%s payment=%s): %s",
            registration_id,
            razorpay_order_id,
            razorpay_payment_id,
            exc,
        )
        registration.payment_status = "failed"
        registration.razorpay_payment_id = razorpay_payment_id
        registration.razorpay_signature = razorpay_signature
        registration.save(
            update_fields=[
                "payment_status",
                "razorpay_payment_id",
                "razorpay_signature",
                "updated_at",
            ]
        )
        _rzp_checkout_session_clear(request)
        return JsonResponse(
            {
                "success": False,
                "error": (
                    "Signature verification failed. Use the same Razorpay Key Id and Key Secret "
                    "as in your Dashboard (test keys with rzp_test_*, live with rzp_live_*)."
                ),
            },
            status=400,
        )
    except Exception as exc:
        logger.exception("Unexpected error during Razorpay verify: %s", exc)
        registration.payment_status = "failed"
        registration.razorpay_payment_id = razorpay_payment_id
        registration.razorpay_signature = razorpay_signature
        registration.save(
            update_fields=[
                "payment_status",
                "razorpay_payment_id",
                "razorpay_signature",
                "updated_at",
            ]
        )
        _rzp_checkout_session_clear(request)
        return JsonResponse({"success": False, "error": "Payment verification error."}, status=500)

    registration.payment_status = "paid"
    registration.razorpay_payment_id = razorpay_payment_id
    registration.razorpay_signature = razorpay_signature
    registration.save(
        update_fields=[
            "payment_status",
            "razorpay_payment_id",
            "razorpay_signature",
            "updated_at",
        ]
    )

    # Keep user record aligned (optional legacy fields used elsewhere)
    if registration.user_id:
        User.objects.filter(id=registration.user_id).update(
            date_time_of_payment=timezone.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

    _rzp_checkout_session_clear(request)
    logger.info(
        "Razorpay verify OK registration_id=%s order_id=%s payment_id=%s",
        registration_id,
        razorpay_order_id,
        razorpay_payment_id,
    )
    return JsonResponse(
        {
            "success": True,
            "redirect_url": reverse("payment_success"),
        }
    )


def payment_success(request):
    return render(request, "success.html")


def payment_failure(request):
    return render(request, "failure.html")


@ensure_csrf_cookie
def registration_payment_redirect(request):
    prefill = request.session.get("payment_prefill")
    if not prefill:
        messages.error(request, "Payment session expired. Start again from registration or your profile.")
        if _portal_session_user(request):
            return redirect("user_profile")
        return redirect("register")

    checkout_bootstrap = None
    pending_id = request.session.get("rzp_pending_registration_id")
    exp = request.session.get("rzp_pending_expires")
    prefill_email = (prefill.get("email") or "").strip().lower()
    prefill_mtype = (prefill.get("membership_type") or "").strip().lower()
    try:
        exp_ts = float(exp) if exp is not None else 0.0
    except (TypeError, ValueError):
        exp_ts = 0.0

    if pending_id is not None and timezone.now().timestamp() < exp_ts:
        reg = (
            PaymentRegistration.objects.filter(pk=pending_id, payment_status="pending")
            .exclude(razorpay_order_id__isnull=True)
            .exclude(razorpay_order_id="")
            .first()
        )
        if reg:
            notes = reg.notes if isinstance(reg.notes, dict) else {}
            reg_m = (notes.get("membership_type") or "").strip().lower()
            if reg.email.strip().lower() == prefill_email and reg_m == prefill_mtype:
                checkout_bootstrap = {
                    "order_id": (reg.razorpay_order_id or "").strip(),
                    "registration_id": reg.pk,
                    "amount": int(reg.amount),
                    "currency": (reg.currency or "INR").strip(),
                    "key": RAZORPAY_KEY_ID,
                }
                logger.info(
                    "Razorpay checkout bootstrap reuse registration_id=%s order_id=%s",
                    reg.pk,
                    checkout_bootstrap["order_id"],
                )
        if checkout_bootstrap is None:
            _rzp_checkout_session_clear(request)
    elif pending_id is not None:
        _rzp_checkout_session_clear(request)

    return render(
        request,
        "registration_payment_redirect.html",
        {
            "prefill": prefill,
            "checkout_bootstrap": checkout_bootstrap,
        },
    )


def register_user_view(request):
    """
    Handle user registration with file uploads.

    This view processes the registration form submission, including file uploads for various
    user documents. It handles the following:
    - Validates the registration form
    - Processes file uploads to MinIO storage
    - Creates a new user account
    - Handles success/failure messages

    Args:
        request: HttpRequest object containing POST data and files

    Returns:
        HttpResponse: Rendered registration page or redirect to login on success

    Raises:
        None - Exceptions are caught and logged
    """
    logger.debug("Starting user registration process")

    selected_membership = request.GET.get("membership_type", "").strip().lower()

    if request.method == "POST":
        logger.info("Processing POST request for user registration")
        selected_membership = request.POST.get("membership_type", "").strip().lower() or selected_membership
        form = UserRegistrationForm(
            request.POST,
            request.FILES,
            membership_type=selected_membership,
        )

        if form.is_valid():
            logger.debug("Registration form is valid")
            user = form.save(commit=False)

            # Prepare file upload paths
            file_fields = [
                ("photo", f"users/{user.username}/photo/"),
                ("state_nmc", f"users/{user.username}/state_nmc/"),
                ("passport", f"users/{user.username}/passport/"),
                (
                    "medical_qualification",
                    f"users/{user.username}/medical_qualification/",
                ),
            ]

            upload_success = True
            uploaded_files = {}
            from django.conf import settings
            import os
            from pathlib import Path

            for field_name, folder_path in file_fields:
                file = request.FILES.get(field_name)
                if file:
                    logger.debug(f"Processing upload for {field_name}")
                    # Save file to local media directory
                    file_name = f"{folder_path}{file.name}"
                    media_path = os.path.join(settings.MEDIA_ROOT, file_name)
                    
                    # Ensure directory exists
                    os.makedirs(os.path.dirname(media_path), exist_ok=True)
                    
                    # Reset file pointer before saving
                    file.seek(0)
                    
                    # Save file to local storage
                    try:
                        with open(media_path, 'wb+') as destination:
                            for chunk in file.chunks():
                                destination.write(chunk)
                        
                        # Store relative path (media/...) for database
                        uploaded_files[field_name] = f"media/{file_name}"
                        logger.info(
                            f"Successfully uploaded {field_name} to local storage: {file_name}"
                        )
                    except Exception as e:
                        error_msg = f"Failed to upload {field_name} to local storage: {str(e)}"
                        logger.error(error_msg)
                        form.add_error(field_name, f"File upload failed: {str(e)}")
                        upload_success = False

            if upload_success:
                # Assign uploaded file paths to the user model fields
                # Ensure required fields have values
                user.photo = uploaded_files.get("photo") or ""
                user.state_nmc = uploaded_files.get("state_nmc") or ""
                user.passport = uploaded_files.get("passport") or ""
                user.medical_qualification = uploaded_files.get("medical_qualification") or ""
                user.payment_transaction_proof = ""
                
                # Validate required fields are not empty
                required_fields = {
                    "photo": user.photo,
                    "passport": user.passport,
                    "medical_qualification": user.medical_qualification,
                }
                
                missing_fields = [field for field, value in required_fields.items() if not value]
                if missing_fields:
                    logger.error(f"Missing required file fields: {missing_fields}")
                    for field in missing_fields:
                        form.add_error(field, f"{field.replace('_', ' ').title()} is required.")
                    upload_success = False
                
                if upload_success:
                    # All files uploaded successfully, proceed with user save
                    user.save()

                    # Map question text to User model field names for backward compatibility
                    question_to_field_map = {
                    'First Name': 'first_name',
                    'Middle Name': 'middle_name',
                    'Last Name': 'last_name',
                    'Gender': 'gender',
                    'WhatsApp Number': 'whatsapp_number',
                    'Mobile Number': 'mobile_number',
                    'Address for Communication': 'address_communication',
                    'Permanent Address': 'address_permanent',
                    'District': 'district',
                    'Father/Spouse Name & Occupation': 'father_spouse_details',
                    'Blood Group': 'blood_group',
                    'Educational Status': 'educational_status',
                    'Category You Belong To For Membership': 'category',
                    'University Name': 'university_name',
                    'Country Where University is situated': 'country_university',
                    'Year of Joining Medical Education': 'year_of_joining',
                    'Year of Completion/Expected Completion of course': 'year_of_completion',
                    'Passport Size Photo': 'photo',
                    'State/NMC registration certificate': 'state_nmc',
                    'Passport (Front and Back)': 'passport',
                    'Medical qualification (Provisional/Permanent) Degree/ Diploma Certificate/ Student ID Card': 'medical_qualification',
                    'Date and time of Payment (Transaction)': 'date_time_of_payment',
                    'Payment Transaction Proof/screenshot': 'payment_transaction_proof',
                    'Willing to be a blood donor?': 'willing_to_be_donor',
                    'Are you in any district group of AKFMA? If yes mention your MID': 'mid',
                    'ADAMS membership number': 'mid',
                    'Agreement for Joining Association of Doctors And Medical Students (ADAMS)': 'agreement',
                    'I agree to the terms and conditions of the association and I hereby submit my application for membership in ADAMS': 'application',
                    }
                    
                    # Save custom question answers and map to User model fields
                    custom_questions = RegistrationQuestion.objects.filter(is_active=True)
                    for question in custom_questions:
                        answer_key = f"custom_question_{question.id}"
                        
                        # Handle file uploads
                        if question.question_type == 'file':
                            uploaded_file = request.FILES.get(answer_key)
                            if uploaded_file:
                                # Validate file type (PDF or JPG)
                                file_ext = uploaded_file.name.split('.')[-1].lower()
                                if file_ext not in ['pdf', 'jpg', 'jpeg', 'png']:
                                    continue  # Skip invalid file types
                                
                                # Upload to local storage
                                folder_path = f"users/{user.username}/custom_questions/"
                                file_name = f"{folder_path}{uploaded_file.name}"
                                media_path = os.path.join(settings.MEDIA_ROOT, file_name)
                                
                                # Ensure directory exists
                                os.makedirs(os.path.dirname(media_path), exist_ok=True)
                                
                                # Reset file pointer
                                uploaded_file.seek(0)
                                
                                try:
                                    # Save file to local storage
                                    with open(media_path, 'wb+') as destination:
                                        for chunk in uploaded_file.chunks():
                                            destination.write(chunk)
                                    
                                    # Store relative path for database
                                    db_file_name = f"media/{file_name}"
                                    answer, created = RegistrationAnswer.objects.get_or_create(
                                        user=user,
                                        question=question,
                                        defaults={}
                                    )
                                    answer.answer_file = db_file_name
                                    answer.save()
                                    logger.info(f"Successfully uploaded custom question file to local storage: {file_name}")
                                except Exception as e:
                                    logger.error(f"Failed to upload custom question file: {e}")
                                    continue  # Skip this file and continue with next
                                    
                                    # Map to User model field if mapping exists
                                    user_field = question_to_field_map.get(question.question_text)
                                    if user_field:
                                        setattr(user, user_field, file_name)
                            elif question.is_required:
                                # Required file but not provided - validation should catch this
                                continue
                        else:
                            # Handle text-based answers
                            answer_value = request.POST.get(answer_key, '').strip()
                            
                            if question.is_required and not answer_value:
                                continue  # Skip if required but empty (form validation should catch this)
                            
                            if answer_value:  # Only save if there's an answer
                                answer, created = RegistrationAnswer.objects.get_or_create(
                                    user=user,
                                    question=question,
                                    defaults={}
                                )
                                
                                # Set answer based on question type
                                if question.question_type == 'date':
                                    try:
                                        from datetime import datetime
                                        answer.answer_date = datetime.strptime(answer_value, '%Y-%m-%d').date()
                                    except:
                                        pass
                                elif question.question_type == 'number':
                                    try:
                                        answer.answer_number = float(answer_value)
                                    except:
                                        pass
                                elif question.question_type in ['checkbox', 'radio']:
                                    answer.answer_boolean = answer_value.lower() in ['true', '1', 'on', 'yes']
                                else:
                                    answer.answer_text = answer_value
                                
                                answer.save()
                                
                                # Map to User model field if mapping exists
                                user_field = question_to_field_map.get(question.question_text)
                                if user_field:
                                    if question.question_type == 'date':
                                        setattr(user, user_field, str(answer.answer_date) if answer.answer_date else '')
                                    elif question.question_type in ['checkbox']:
                                        setattr(user, user_field, answer.answer_boolean)
                                    else:
                                        setattr(user, user_field, answer.answer_text or '')
                    
                    # Save user after mapping all fields (if any custom questions updated fields)
                    user.save()

                    logger.info(
                        f"New user registered successfully - username: {form.cleaned_data['username']}"
                    )

                    plain_password = (form.cleaned_data.get("password1") or "").strip()
                    try:
                        ok_mail = send_new_account_credentials_email(
                            username=user.username,
                            email=user.email,
                            password_plain=plain_password,
                            login_url=request.build_absolute_uri(reverse("login")),
                            profile_url=request.build_absolute_uri(reverse("user_profile")),
                            upload_url=request.build_absolute_uri(
                                reverse("member_complete_uploads")
                            ),
                        )
                        if not ok_mail:
                            logger.warning(
                                "Welcome email was not sent for new user %s",
                                user.email,
                            )
                    except Exception:
                        logger.exception(
                            "Welcome email failed for new user %s",
                            user.email,
                        )

                    request.session["payment_prefill"] = {
                        "name": f"{user.first_name} {user.last_name}".strip() or user.username,
                        "email": user.email,
                        "phone": user.mobile_number,
                        "membership_type": selected_membership,
                        "membership_label": _get_membership_details(selected_membership)["label"] if _get_membership_details(selected_membership) else "",
                        "amount": _get_membership_details(selected_membership)["amount"] if _get_membership_details(selected_membership) else MEMBERSHIP_TYPES["adhoc"]["amount"],
                    }
                    messages.success(request, "Registration completed. Redirecting to payment...")
                    return redirect("registration_payment_redirect")
            else:
                logger.error(
                    f"Registration failed due to file upload errors: {form.errors}"
                )
        else:
            logger.error(f"Registration form validation failed: {form.errors}")
            messages.error(
                request, "There was an error with your registration. Please try again."
            )
    else:
        logger.debug("Displaying empty registration form")
        initial_data = {}
        form = UserRegistrationForm(
            initial=initial_data,
            membership_type=selected_membership,
        )
    
    # Get active custom questions for registration form
    custom_questions = RegistrationQuestion.objects.filter(is_active=True).order_by('order', 'created_at')

    return render(request, "register.html", {
        "form": form,
        "custom_questions": custom_questions,
        "selected_membership": selected_membership,
        "selected_membership_details": _get_membership_details(selected_membership),
    })


def update_user_view(request, user_id):
    """
    View function to handle user profile updates including file uploads.

    This view allows updating user profile information and associated files like photos,
    documents etc. It handles both GET and POST requests:
    - GET: Displays the update form with current user data
    - POST: Processes the form submission and file uploads

    Args:
        request (HttpRequest): The request object
        user_id (int): ID of the user to be updated

    Returns:
        HttpResponse: Rendered update form on GET or redirect on successful POST

    Raises:
        User.DoesNotExist: If user with given ID is not found
    """
    logger.info(f"Accessing update view for user ID: {user_id}")
    user = User.objects.get(id=user_id)
    logger.debug(f"Retrieved user: {user.username}")

    if request.method == "POST":
        logger.info(f"Processing POST request to update user {user.username}")
        form = UserUpdateForm(request.POST, request.FILES, instance=user)

        # Get existing file paths
        existing_photo = getattr(user, "photo")
        existing_state_ncm = getattr(user, "state_nmc")
        existing_passport = getattr(user, "passport")
        existing_medical_qualification = getattr(user, "medical_qualification")
        existing_payment_transaction_proof = getattr(user, "payment_transaction_proof")

        existing_files = {
            "photo": existing_photo,
            "state_nmc": existing_state_ncm,
            "passport": existing_passport,
            "medical_qualification": existing_medical_qualification,
            "payment_transaction_proof": existing_payment_transaction_proof,
        }

        if form.is_valid():
            logger.debug("Form validation successful")
            # Prepare file upload paths
            file_fields = [
                ("photo", f"users/{user.username}/photo/"),
                ("state_nmc", f"users/{user.username}/state_nmc/"),
                ("passport", f"users/{user.username}/passport/"),
                (
                    "medical_qualification",
                    f"users/{user.username}/medical_qualification/",
                ),
                (
                    "payment_transaction_proof",
                    f"users/{user.username}/payment_transaction_proof/",
                ),
            ]

            # Track uploaded files
            uploaded_files = {}
            upload_success = True
            # Process file fields to upload and preserve existing files
            for field_name, folder_path in file_fields:
                file = request.FILES.get(field_name)
                if file:
                    logger.debug(f"Processing file upload for {field_name}")
                    # If a file is uploaded, handle the upload process
                    file_name = f"{folder_path}{file.name}"
                    media_path = os.path.join(settings.MEDIA_ROOT, file_name)
                    
                    # Ensure directory exists
                    os.makedirs(os.path.dirname(media_path), exist_ok=True)
                    
                    # Reset file pointer
                    file.seek(0)
                    
                    try:
                        # Save file to local storage
                        with open(media_path, 'wb+') as destination:
                            for chunk in file.chunks():
                                destination.write(chunk)
                        
                        # Store relative path for database
                        uploaded_files[field_name] = f"media/{file_name}"
                        logger.info(
                            f"Successfully uploaded {field_name} to local storage: {file_name}"
                        )
                    except Exception as e:
                        logger.error(f"Failed to upload {field_name}: {e}")
                        form.add_error(field_name, f"File upload failed: {str(e)}")
                        upload_success = False
                    else:
                        form.add_error(field_name, f"Failed to upload {field_name}.")
                        logger.error(f"Failed to upload {field_name}")
                        upload_success = False
                else:
                    # If no file is uploaded, keep the existing file
                    uploaded_files[field_name] = existing_files.get(field_name)
                    logger.debug(f"Keeping existing file for {field_name}")

            if upload_success:
                # Save the updated user object with preserved files if necessary
                updated_user = form.save(commit=False)

                for field_name in uploaded_files:
                    # Assign uploaded or existing file paths
                    setattr(updated_user, field_name, uploaded_files[field_name])

                updated_user.save()

                logger.info(f"Successfully updated profile for user {user.username}")
                messages.success(request, "Profile updated successfully.")
                send_email_with_attachments(
                    subject="Profile Update Confirmation",
                    template_type="custom",
                    custom_message="Your profile has been updated successfully. If you did not make this change, please contact support immediately.",
                    recipient_list=[user.email],
                )
                return redirect("admin_panel")
        else:
            logger.error(f"Profile update failed: {form.errors}")
            messages.error(
                request, "There was an error updating your profile. Please try again."
            )
    else:
        logger.debug(f"Displaying update form for user {user.username}")
        form = UserUpdateForm(instance=user)

    return render(request, "update_profile.html", {"form": form})


def delete_user_view(request, user_id):
    """
    View function to handle user deletion including associated files.

    This view deletes a user and their associated files (photos, documents etc) from both
    the database and MinIO storage. It handles both GET and POST requests:
    - GET: Shows confirmation page
    - POST: Performs the actual deletion

    Args:
        request (HttpRequest): The request object
        user_id (int): ID of the user to be deleted

    Returns:
        HttpResponse: Rendered confirmation page on GET or redirect on successful POST

    Raises:
        Http404: If user with given ID is not found
    """
    logger.info(f"Accessing delete view for user ID: {user_id}")
    user = get_object_or_404(User, id=user_id)
    logger.debug(f"Retrieved user: {user.username}")

    if request.method == "POST":
        logger.info(f"Processing POST request to delete user {user.username}")
        # Delete associated files from MinIO
        file_fields = [
            "photo",
            "state_nmc",
            "passport",
            "medical_qualification",
            "payment_transaction_proof",
        ]

        for field in file_fields:
            file_path = getattr(user, field)
            if file_path:
                logger.debug(f"Attempting to delete {field} file: {file_path}")
                try:
                    # Delete from local storage
                    pdf_path = str(file_path)
                    if not pdf_path.startswith("media/"):
                        pdf_path = f"media/{pdf_path}"
                    
                    full_path = os.path.join(settings.BASE_DIR, pdf_path)
                    full_path = os.path.normpath(full_path)
                    
                    # Security check
                    media_root = os.path.normpath(settings.MEDIA_ROOT)
                    if full_path.startswith(media_root) and os.path.exists(full_path):
                        os.remove(full_path)
                        logger.debug(f"Successfully deleted {field} file")
                    else:
                        logger.warning(f"File not found or outside media root: {full_path}")
                except Exception as e:
                    logger.error(f"Failed to delete {field} file for user {user.username}: {e}")
                    messages.error(request, f"Failed to delete {field}.")
                    return redirect("admin_panel")

        # Delete the user instance
        try:
            user.delete()
            logger.info(
                f"Successfully deleted user {user.username} and associated files"
            )
            messages.success(request, f"User {user.username} deleted successfully.")
        except Exception as e:
            logger.error(f"Error deleting user {user.username}: {str(e)}")
            messages.error(
                request, f"An error occurred while deleting the user: {str(e)}"
            )
            return redirect("admin_panel")

        return redirect("admin_panel")

    logger.debug(f"Displaying delete confirmation page for user {user.username}")
    return render(request, "delete_user.html", {"user": user})


def login_view(request):
    """
    Handle user login authentication and OTP verification flow.

    This view processes login form submissions and manages the login flow:
    - Validates login credentials
    - Sends OTP for verification
    - Manages user session
    - Handles error cases

    Args:
        request: HttpRequest object containing POST data or session info

    Returns:
        HttpResponse: Rendered login page or redirect to OTP verification

    Logs:
        - Info: Successful login attempts and OTP sending
        - Warning: Failed login attempts
        - Error: Form validation errors
        - Debug: Session username checks
    """
    logger.debug("Processing login view request")
    user = None

    if request.method == "POST":
        logger.info("Processing POST login request")
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            logger.debug("Login form is valid, attempting authentication")
            user = authenticate(
                request,
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password"],
            )
            if user is not None:
                logger.info(f"User {user.username} authenticated successfully")
                request.session["user_id"] = user.id
                request.session["username"] = user.username
                
                # Check if 2FA is enabled
                two_factor_enabled = SystemSettings.get_setting('two_factor_enabled', default='true')
                two_factor_enabled = str(two_factor_enabled).lower() in ['true', '1', 'yes', 'on']
                
                if two_factor_enabled:
                    # Send OTP with error handling
                    try:
                        otp_sent = send_otp_via_email(user)
                        if otp_sent:
                            logger.info(f"OTP sent to user {user.username}'s email: {user.email}")
                            messages.success(
                                request, f"OTP sent to your registered email ({user.email}). Please check your inbox and spam folder."
                            )
                        else:
                            logger.error(f"Failed to send OTP to user {user.username}")
                            messages.error(
                                request, "Failed to send OTP email. Please try again or contact support."
                            )
                            return redirect("login")
                    except Exception as e:
                        logger.exception(f"Error sending OTP to user {user.username}: {e}")
                        messages.error(
                            request, f"Error sending OTP: {str(e)}. Please try again."
                        )
                        return redirect("login")
                    
                    return redirect("verify_otp")
                else:
                    # 2FA is disabled, login directly
                    login(request, user)
                    logger.info(f"User {user.username} logged in directly (2FA disabled)")
                    messages.success(request, "Login successful.")
                    return redirect("home")
            else:
                logger.warning(
                    f"Failed login attempt for username: {form.cleaned_data['username']}"
                )
                messages.error(request, "Invalid username or password.")
        else:
            logger.error(f"Login form errors: {form.errors}")
            messages.error(request, "Invalid username or password.")
    else:
        uname = request.session.get("username")
        if uname:
            logger.debug(f"Found username {uname} in session")
            try:
                user = User.objects.get(username=uname)
                logger.debug(f"Retrieved user object for {uname}")
            except User.DoesNotExist:
                logger.warning(f"User {uname} from session not found in database")
                user = None
        form = UserLoginForm()

    data = {"form": form, "user": user}
    logger.debug("Rendering login page")
    return render(request, "login.html", {"data": data})


def user_logout_view(request):
    """
    Handle user logout and redirect to login page.

    This view logs out the current user by clearing their session data and
    authentication status. It performs the following:
    - Logs out the user using Django's logout() function
    - Displays a success message
    - Redirects to the login page

    Args:
        request: HttpRequest object containing session and auth data

    Returns:
        HttpResponseRedirect: Redirect to login page

    Logs:
        - Info: When user successfully logs out
    """
    logger.info(f"User {request.user} logging out")
    logout(request)
    messages.success(request, "Logged out successfully.")
    logger.info("User logged out successfully")
    return redirect("login")


def home_page_view(request):
    """
    Render the home page with user and announcement data.

    This view handles displaying the home page by:
    - Retrieving the current user from session if logged in
    - Fetching all announcements from the database
    - Rendering the home template with user and announcement data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered home.html template with user and announcement data

    Logs:
        - Info: When user accesses home page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing home page request")

    # Get username from session
    uname = request.session.get("username")

    user = None
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the home page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    # Fetch announcements based on user authentication
    try:
        if user:
            # Logged-in users see all announcements
            announcements = Announcement.objects.all()
        else:
            # Public visitors only see public announcements
            # Handle cases where visibility field might not exist yet
            announcements = Announcement.objects.filter(visibility='public') | Announcement.objects.filter(visibility__isnull=True)
    except Exception as e:
        # Fallback if visibility field doesn't exist
        logger.warning(f"Error filtering by visibility: {str(e)}")
        announcements = Announcement.objects.all()
    
    logger.debug(f"Retrieved {len(announcements)} announcements")

    # Prepare context data
    data = {
        "user": user,
        "announcements": announcements,
    }
    logger.debug("Rendering home page")
    return render(request, "home.html", {"data": data})


@login_required(login_url="/login/")
@admin_required
def add_announcement(request):
    """
    Add a new announcement to the system.

    This view handles the creation of new announcements by admin users. It performs the following:
    - Validates that the user is an admin
    - Creates a new announcement with the provided text and link
    - Logs the announcement creation
    - Handles validation and permission errors

    Args:
        request: HttpRequest object containing POST data and session info

    Returns:
        HttpResponseRedirect: Redirect to home page after announcement creation

    Logs:
        - Info: When announcement is successfully created
        - Warning: When empty announcement attempted or non-admin tries to create
    """
    logger.debug("Processing add announcement request")

    if request.method == "POST":
        uname = request.session.get("username")
        logger.debug(f"Getting user for username: {uname}")

        user = User.objects.get(username=uname)
        if user.role == "admin":  # Check if the user is an admin
            announcement_text = request.POST.get("announcement")
            announcement_link = request.POST.get("announcement_link", "/")

            logger.debug(f"Received announcement text: {announcement_text}")
            logger.debug(f"Received announcement link: {announcement_link}")

            if announcement_text:
                # Create the announcement
                Announcement.objects.create(
                    uid=user,
                    announcement=announcement_text,
                    hyper_link=announcement_link,
                )
                logger.info(f"Announcement added by admin {user.username}")
                messages.success(request, "Announcement added successfully.")
            else:
                messages.error(request, "Announcement text cannot be empty.")
                logger.warning(
                    f"Attempted to add an empty announcement by admin {user.username}"
                )
        else:
            logger.warning(
                f"User {user.username} attempted to add an announcement without admin role."
            )
            messages.error(
                request,
                "You do not have the required permissions to add an announcement.",
            )

    logger.debug("Redirecting to home page")
    return redirect("home")


@login_required(login_url="/login/")
@admin_required
def delete_announcement(request, announcement_id):
    """
    Delete an announcement from the system.

    This view handles the deletion of announcements by admin users. It performs the following:
    - Validates that the user is an admin
    - Deletes the specified announcement
    - Logs the deletion action
    - Handles validation and permission errors

    Args:
        request: HttpRequest object containing session info
        announcement_id: ID of the announcement to delete

    Returns:
        HttpResponseRedirect: Redirect to home page after announcement deletion

    Logs:
        - Info: When announcement is successfully deleted
        - Warning: When non-admin tries to delete announcement
        - Debug: When delete request is received
    """
    logger.debug(f"Processing delete announcement request for ID: {announcement_id}")

    if request.method == "POST":
        uname = request.session.get("username")
        logger.debug(f"Getting user for username: {uname}")

        user = User.objects.get(username=uname)
        if user.role == "admin":  # Check if the user is an admin
            announcement = get_object_or_404(Announcement, aid=announcement_id)
            announcement.delete()
            logger.info(
                f"Announcement {announcement_id} deleted by admin {user.username}"
            )
            messages.success(request, "Announcement deleted successfully.")
        else:
            logger.warning(
                f"User {user.username} attempted to delete an announcement without admin role."
            )
            messages.error(
                request,
                "You do not have the required permissions to delete an announcement.",
            )

    logger.debug("Redirecting to home page")
    return redirect("home")


def render_gallery(request):
    """
    Render the gallery page with user info and images.

    This view handles displaying the gallery page by:
    - Retrieving the current user from session if logged in
    - Fetching all gallery images from storage
    - Rendering the gallery template with user and image data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered gallery.html template with user and image data

    Logs:
        - Info: When user accesses gallery page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing gallery page request")

    uname = request.session.get("username")
    user = None

    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the gallery page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    images = get_gallery_images()
    logger.debug(f"Retrieved {len(images)} gallery images")

    data = {"user": user, "images": images}
    logger.debug("Rendering gallery page")
    return render(request, "gallery.html", {"data": data})


def serve_image(request, image_name):
    """
    Serve an image from MinIO storage.

    This view retrieves an image from MinIO storage and returns it as an HTTP response.
    The image is served inline for display in the browser.

    Args:
        request: HttpRequest object
        image_name (str): Name/path of the image file in MinIO storage

    Returns:
        HttpResponse: Image file with appropriate content type headers

    Raises:
        S3Error: If there is an error retrieving the image from MinIO
    """
    logger.debug(f"Attempting to serve image: {image_name}")
    try:
        # Preferred: serve from MinIO when enabled
        if getattr(settings, "MINIO_ENABLED", False):
            from .utils import fetch_file_from_minio

            image_data = fetch_file_from_minio(image_name)
            if image_data is None:
                logger.error(f"MinIO image not found or fetch failed: {image_name}")
                return HttpResponse("Image not found", status=404)

            if image_name.lower().endswith(".png"):
                content_type = "image/png"
            elif image_name.lower().endswith((".jpg", ".jpeg")):
                content_type = "image/jpeg"
            elif image_name.lower().endswith(".webp"):
                content_type = "image/webp"
            else:
                content_type = "application/octet-stream"

            response = HttpResponse(image_data, content_type=content_type)
            response["Content-Disposition"] = f"inline; filename={os.path.basename(image_name)}"
            return response

        # Fallback: local storage (backward compatibility)
        if not image_name.startswith("media/"):
            image_name = f"media/{image_name}"

        full_path = os.path.join(settings.BASE_DIR, image_name)
        full_path = os.path.normpath(full_path)

        # Security check
        media_root = os.path.normpath(settings.MEDIA_ROOT)
        if not full_path.startswith(media_root):
            logger.error(f"Security violation: File path outside MEDIA_ROOT: {image_name}")
            return HttpResponse(status=403, content="Access denied")

        if not os.path.exists(full_path):
            logger.error(f"Image not found: {full_path}")
            return HttpResponse("Image not found", status=404)

        with open(full_path, "rb") as f:
            image_data = f.read()

        if image_name.lower().endswith(".png"):
            content_type = "image/png"
        elif image_name.lower().endswith((".jpg", ".jpeg")):
            content_type = "image/jpeg"
        elif image_name.lower().endswith(".webp"):
            content_type = "image/webp"
        else:
            content_type = "application/octet-stream"

        response = HttpResponse(image_data, content_type=content_type)
        response["Content-Disposition"] = f"inline; filename={os.path.basename(image_name)}"
        return response

    except Exception as e:
        logger.exception(f"Error serving image {image_name}: {str(e)}")
        return HttpResponse(f"Error serving image: {e}", status=500)


@login_required(login_url="/login/")
@admin_required
def upload_image_view(request):
    """
    Handle image upload to MinIO storage and save metadata to database.

    This view processes image file uploads from admin users and stores them in MinIO.
    It performs validation on the file type and logs the upload process.

    Args:
        request: HttpRequest object containing POST data and files

    Returns:
        HttpResponseRedirect: Redirect to gallery page after upload attempt

    Logs:
        - Info: When file is received and successfully uploaded
        - Warning: When invalid file type is attempted
        - Error: When upload fails
    """
    if request.method == "POST" and request.FILES.get("image"):
        from .models import GalleryImage
        
        image = request.FILES["image"]
        title = request.POST.get("title", "").strip()[:200]
        description = request.POST.get("description", "").strip()
        file_name = image.name
        file_size = image.size
        file_type = image.content_type

        logger.info(
            f"Received file: {file_name}, size: {file_size} bytes, type: {file_type}"
        )

        if file_type.startswith("image/"):
            success = False
            db_file_name = None
            local_media_path = None

            # Preferred: upload to MinIO when enabled
            if getattr(settings, "MINIO_ENABLED", False):
                from uuid import uuid4
                from django.utils.text import get_valid_filename
                from .utils import upload_file_to_minio

                safe_name = get_valid_filename(file_name) or "image"
                object_name = f"gallery-{uuid4().hex}-{safe_name}"

                success = upload_file_to_minio(image, object_name, content_type=file_type)
                if success:
                    db_file_name = object_name
                    logger.info(f"Successfully uploaded gallery image to MinIO: {object_name}")
                else:
                    logger.error(f"Failed to upload gallery image to MinIO: {object_name}")
            else:
                # Local fallback: unique object name (image_name is unique in DB)
                from uuid import uuid4
                from django.utils.text import get_valid_filename

                safe_name = get_valid_filename(file_name) or "image"
                relative_name = f"gallery-{uuid4().hex}-{safe_name}"
                local_media_path = os.path.join(settings.MEDIA_ROOT, relative_name)
                os.makedirs(os.path.dirname(local_media_path), exist_ok=True)
                image.seek(0)
                try:
                    with open(local_media_path, "wb+") as destination:
                        for chunk in image.chunks():
                            destination.write(chunk)
                    db_file_name = f"media/{relative_name}"
                    success = True
                    logger.info(
                        f"Successfully uploaded gallery image to local storage: {relative_name}"
                    )
                except Exception as e:
                    logger.error(f"Failed to upload gallery image locally: {e}")
                    success = False

            if success:
                # Save to database
                try:
                    user = request.user if request.user.is_authenticated else None
                    GalleryImage.objects.create(
                        image_name=db_file_name,
                        title=title or None,
                        description=description or None,
                        uploaded_by=user
                    )
                    messages.success(request, f"Image '{file_name}' uploaded successfully.")
                    logger.info(f"Image '{file_name}' uploaded and saved to database as '{db_file_name}'.")
                except Exception as e:
                    logger.exception(f"Error saving image metadata to database: {e}")
                    if local_media_path and os.path.isfile(local_media_path):
                        try:
                            os.remove(local_media_path)
                        except OSError:
                            pass
                    messages.warning(
                        request,
                        "Could not save image details after upload. Please try again or contact support.",
                    )
            else:
                messages.error(request, "Failed to upload the image. Please try again.")
                logger.error(f"Failed to upload image '{file_name}'.")
        else:
            messages.error(request, "Invalid file type. Please upload a valid image.")
            logger.warning(
                f"Invalid file type attempted: {file_type} for file '{file_name}'."
            )

    return redirect("gallery")


@login_required(login_url="/login/")
@admin_required
def delete_image_view(request, image_name):
    """
    Handle deletion of images from MinIO storage by admin users.

    This view processes image deletion requests from admin users. It validates the request
    and attempts to delete the specified image from MinIO storage.

    Args:
        request: HttpRequest object containing session info and POST data
        image_name (str): Name/path of the image file to delete from MinIO

    Returns:
        HttpResponseRedirect: Redirect to gallery page after deletion attempt

    Logs:
        - Info: When image is successfully deleted
        - Error: When deletion fails
        - Debug: When delete request is received
    """
    logger.debug(f"Processing delete request for image: {image_name}")

    if request.method == "POST":
        try:
            # Preferred: delete from MinIO when enabled
            if getattr(settings, "MINIO_ENABLED", False):
                from .utils import delete_file_from_minio

                if delete_file_from_minio(image_name):
                    messages.success(request, f"Image '{image_name}' deleted successfully.")
                    logger.info(f"Image '{image_name}' deleted from MinIO.")
                else:
                    messages.error(request, f"Failed to delete image '{image_name}'. Please try again.")
                    logger.error(f"Failed to delete image '{image_name}' from MinIO.")
            else:
                # Local fallback (backward compatibility)
                if not image_name.startswith("media/"):
                    image_name = f"media/{image_name}"

                full_path = os.path.join(settings.BASE_DIR, image_name)
                full_path = os.path.normpath(full_path)

                media_root = os.path.normpath(settings.MEDIA_ROOT)
                if full_path.startswith(media_root) and os.path.exists(full_path):
                    os.remove(full_path)
                    messages.success(request, f"Image '{image_name}' deleted successfully.")
                    logger.info(f"Image '{image_name}' deleted from local storage.")
                else:
                    messages.error(request, f"Image file not found: '{image_name}'.")
                    logger.error(f"Image file not found: '{image_name}'")
        except Exception as e:
            messages.error(request, f"Failed to delete image '{image_name}'. Please try again.")
            logger.error(f"Failed to delete image '{image_name}': {e}")

    logger.debug("Redirecting to gallery page")
    return redirect("gallery")


@login_required(login_url="/login/")
def generate_qr_code(request, user_id):
    """
    Generate a QR code for a specific user and return it as a downloadable PNG file.

    This view generates a QR code containing user information and returns it as an HTTP response.
    The QR code is generated dynamically using the user's data.

    Args:
        request: HttpRequest object
        user_id (int): ID of the user to generate QR code for

    Returns:
        HttpResponse: PNG image file containing the generated QR code

    Raises:
        Http404: If user with given ID is not found
    """
    logger.debug(f"Generating QR code for user ID: {user_id}")

    user = get_object_or_404(User, id=user_id)
    logger.info(f"Retrieved user {user.username} for QR code generation")

    try:
        buffer = generate_qr_code_live(user)
        logger.info(f"Successfully generated QR code for user {user.username}")

        response = HttpResponse(buffer, content_type="image/png")
        response["Content-Disposition"] = (
            f'attachment; filename="{user.username}_qr.png"'
        )
        return response

    except Exception as e:
        logger.error(f"Error generating QR code for user {user.username}: {str(e)}")
        raise


@login_required(login_url="/login/")
@admin_required
def render_qr_scanner(request):
    """
    Render the QR code scanner page for admin users.

    This view handles displaying the QR code scanner interface by:
    - Retrieving the current user from session if logged in
    - Validating admin permissions
    - Rendering the QR scanner template with user data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered verify_qr.html template with user data

    Logs:
        - Info: When admin user accesses QR scanner page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing QR scanner page request")

    uname = request.session.get("username")
    # Initialize user and role variables in case user is not logged in
    user = None

    # Check if username exists in session and fetch user and role
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"Admin user {uname} accessed the QR scanner page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    data = {"user": user}
    logger.debug("Rendering QR scanner page")
    return render(request, "verify_qr.html", {"data": data})


@login_required(login_url="/login/")
@admin_required
@csrf_exempt  # For testing; replace with CSRF token handling in production
def validate_qr_code(request):
    """
    Validate QR code data and return user information.

    This view validates QR code data containing a username and returns the corresponding user's
    information along with their photo if available. The view performs the following:
    - Validates that the request is a POST request
    - Decodes the username from the QR code data
    - Checks if the user exists and their account status
    - Retrieves and encodes the user's photo if available
    - Returns user details and photo data in JSON format

    Args:
        request: HttpRequest object containing POST data with QR code information

    Returns:
        JsonResponse: User details and photo data on success, error message on failure

    Raises:
        User.DoesNotExist: If user with given username is not found
        Exception: For any other errors during processing

    Status Codes:
        200: Success
        400: Invalid request or user not found
        403: Account pending/rejected
        404: User or photo not found
        405: Invalid request method
    """
    logger.debug("Processing QR code validation request")

    if request.method == "POST":
        try:
            # Decode username from QR code data
            username = request.body.decode("utf-8").strip('"')
            logger.info(f"Attempting to validate QR code for username: {username}")

            try:
                user = User.objects.get(username=username)
                logger.debug(f"Found user: {username}")

                # Check user account status
                if user.status == "pending":
                    logger.warning(f"Account pending for user: {username}")
                    return JsonResponse(
                        {
                            "status": "failed",
                            "message": "This account is pending approval",
                        },
                        status=403,
                    )
                elif user.status == "rejected":
                    logger.warning(f"Account rejected for user: {username}")
                    return JsonResponse(
                        {
                            "status": "failed",
                            "message": "This account has been rejected",
                        },
                        status=403,
                    )

            except User.DoesNotExist:
                logger.error(f"User not found: {username}")
                return JsonResponse(
                    {"status": "failed", "message": "User not found"}, status=404
                )

            # Process user photo if available
            photo_data = None
            if user.photo:
                logger.debug(f"Fetching photo for user: {username}")
                image_path = str(user.photo)
                
                # If path doesn't start with "media/", add it
                if not image_path.startswith("media/"):
                    image_path = f"media/{image_path}"
                
                full_path = os.path.join(settings.BASE_DIR, image_path)
                full_path = os.path.normpath(full_path)
                
                # Security check
                media_root = os.path.normpath(settings.MEDIA_ROOT)
                if not full_path.startswith(media_root):
                    logger.error(f"Security violation: File path outside MEDIA_ROOT: {image_path}")
                    return JsonResponse({"error": "Access denied"}, status=403)
                
                try:
                    if os.path.exists(full_path):
                        with open(full_path, 'rb') as f:
                            image_data = f.read()
                        photo_data = base64.b64encode(image_data).decode("utf-8")
                        logger.debug(f"Successfully encoded photo for user: {username}")
                    else:
                        logger.error(f"Photo not found for user: {username} at {full_path}")
                        return JsonResponse({"error": "Image not found"}, status=404)
                except Exception as e:
                    logger.error(f"Error reading photo for user {username}: {e}")
                    return JsonResponse({"error": "Error reading image"}, status=500)

            # Return user data
            logger.info(f"Successfully validated QR code for user: {username}")
            return JsonResponse(
                {
                    "status": "success",
                    "user": {
                        "username": user.username,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "email": user.email,
                        "gender": user.gender,
                        "mobile_number": user.mobile_number,
                        "address_communication": user.address_communication,
                        "district": user.district,
                        "blood_group": user.blood_group,
                        "role": user.role,
                    },
                    "photo": photo_data,
                },
                safe=False,
            )

        except User.DoesNotExist:
            logger.error("User not found in database")
            return JsonResponse(
                {"status": "failed", "message": "User not found"}, status=400
            )
        except Exception as e:
            logger.error(f"Error validating QR code: {str(e)}")
            return JsonResponse({"status": "failed", "message": str(e)}, status=400)

    logger.warning("Invalid request method for QR code validation")
    return JsonResponse(
        {"status": "failed", "message": "Invalid request method"}, status=405
    )


def render_maintanance_page(request):
    """
    Render the maintenance page with user data.

    This view handles displaying the maintenance page by:
    - Retrieving the current user from session if logged in
    - Rendering the inner-page template with user data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered inner-page.html template with user data

    Logs:
        - Info: When user accesses maintenance page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing maintenance page request")

    uname = request.session.get("username")
    user = None
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the maintenance page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    data = {
        "user": user,
    }
    logger.debug("Rendering maintenance page")
    return render(request, "inner-page.html", {"data": data})


def forgot_password(request):
    """
    Handle password reset requests by generating and emailing a new password.

    This view processes password reset requests by:
    - Validating the provided email address
    - Generating a new random password
    - Sending the new password via email
    - Updating the user's password in the database

    Args:
        request: HttpRequest object containing POST data with email

    Returns:
        HttpResponseRedirect: Redirect to login page with success/error message

    Logs:
        - Info: When password reset is requested and completed successfully
        - Warning: When email is not found or invalid request
        - Error: When email sending fails
    """
    logger.debug("Processing password reset request")

    if request.method == "POST":
        email = request.POST.get("email")
        logger.info(f"Password reset requested for email: {email}")

        if not email:
            logger.warning("Email address not provided in password reset request")
            messages.error(request, "Email address is required.")
            return redirect("login")

        try:
            user = User.objects.filter(email=email).first()
            if not user:
                logger.warning(f"No user found for email: {email}")
                messages.error(request, "No user found with this email address.")
                return redirect("login")

            new_password = generate_random_password()
            user.set_password(new_password)
            user.save()
            logger.info(f"New password generated and saved for user: {user.username}")

            try:
                html_content = f"""<html>
                        <body style="font-family: 'Arial', sans-serif; background-color: #f9f9f9; margin: 0; padding: 20px;">
                            <table width="100%" style="max-width: 600px; margin: auto; background-color: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);">
                                <tr>
                                    <td style="padding: 20px; text-align: center; background-color: #2c3e50; border-top-left-radius: 10px; border-top-right-radius: 10px;">
                                        <h1 style="color: #ffffff; margin: 0;">Association of Doctors and Medical Students (ADAMS)</h1>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 20px;">
                                        <h2 style="color: #2c3e50;">🔐 Password Reset Request</h2>
                                        <p style="color: #333333;">
                                            Hello {user.first_name},  
                                        </p>
                                        <p style="color: #555555;">
                                            We have received a request to reset the password for your ADAMS account. Below is your new password:
                                        </p>
                                        <div style="background-color: #f1f8e9; padding: 15px; text-align: center; border-radius: 5px; margin: 20px 0;">
                                            <span style="font-size: 24px; font-weight: bold; color: #28a745;">{new_password}</span>
                                        </div>
                                        <p style="color: #555555;">
                                            Please use this password to log in. <strong>We strongly recommend changing your password immediately</strong> after logging in to keep your account secure.
                                        </p>
                                        <p style="color: #e74c3c;">
                                            ⚠️ If you did not request this password reset, please contact our support team immediately to secure your account.
                                        </p>
                                        <p style="color: #7f8c8d;">Warm regards, <br>The ADAMS Support Team</p>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="background-color: #ecf0f1; padding: 10px; text-align: center; font-size: 14px; border-bottom-left-radius: 10px; border-bottom-right-radius: 10px;">
                                        <p style="margin: 0;">Need help? <a href="https://adams.org.in/contact" style="color: #3498db;">Contact Us</a></p>
                                    </td>
                                </tr>
                            </table>
                        </body>
                    </html>
                    """
                subject="Password Reset Request"
                recipient_list=email,
                # Create email message
                email = EmailMessage(
                    subject=subject,
                    body=html_content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=recipient_list,
                )

                # Set HTML content type
                email.content_subtype = "html"

                # Send email
                email.send(fail_silently=False)

                logger.info(f"Password reset email sent successfully to: {email}")
                messages.success(request, "A new password has been sent to your email.")
                return redirect("login")
            except Exception as e:
                logger.error(f"Error sending password reset email: {str(e)}")
                messages.error(
                    request,
                    "There was an issue sending the email. Please try again later.",
                )
                return redirect("login")

        except User.DoesNotExist:
            logger.warning(f"No user found for email: {email}")
            messages.error(request, "No user found with this email address.")
            return redirect("login")

    logger.warning("Invalid request method for password reset")
    messages.error(request, "Invalid request method.")
    return redirect("login")


def verify_otp_view(request):
    """
    Handle OTP verification for user login.

    This view processes OTP verification by:
    - Validating the submitted OTP code
    - Checking if the OTP matches and is still valid
    - Logging the user in if verification succeeds

    Args:
        request: HttpRequest object containing POST data with OTP code

    Returns:
        HttpResponse: Rendered verify_otp.html template or redirect to home/login

    Logs:
        - Info: OTP verification attempts and successful logins
        - Warning: Invalid sessions and expired OTPs
        - Error: Invalid OTPs and user not found errors
    """
    if request.method == "POST":
        otp_code = "".join(request.POST.getlist("otp_code[]"))  # Combine OTP cells
        user_id = request.session.get("user_id")

        # Log OTP and user data for debugging purposes
        logger.info(
            f"OTP verification request received. User ID from session: {user_id}, OTP code: {otp_code}"
        )

        if not user_id:
            logger.warning("User ID is not present in session.")
            messages.error(request, "Session expired or invalid.")
            return redirect("login")

        try:
            user = User.objects.get(id=user_id)
            logger.info(f"User found: {user.username}")

            otp_record = OTP.objects.get(user=user, otp_code=otp_code)
            logger.info(f"OTP record found for user {user.username}")

            if otp_record.is_valid():
                login(request, user)
                request.session["user_id"] = user.id
                request.session["username"] = user.username
                messages.success(request, "Login successful.")
                logger.info(f"User {user.username} logged in successfully.")
                return redirect("home")
            else:
                messages.error(request, "OTP has expired. Please try again.")
                logger.warning(f"OTP expired for user {user.username}.")
        except (User.DoesNotExist, OTP.DoesNotExist):
            messages.error(request, "Invalid OTP.")
            logger.error(f"Invalid OTP or user not found for user ID {user_id}.")

    return render(request, "verify_otp.html")


@csrf_exempt  # Disable CSRF protection for the view (not recommended in production)
@require_POST  # Ensure it's a POST request
def resend_otp(request):
    """
    Handle OTP resend requests.

    This view processes requests to resend OTP codes by:
    - Validating the request contains proper JSON data
    - Checking if user exists in session
    - Resending OTP via email if request is valid

    Args:
        request: HttpRequest object containing POST data with resend action

    Returns:
        JsonResponse: Success/failure status and message

    Status Codes:
        200: Success
        400: Invalid request data or session expired
        404: User not found
        500: Server error

    Logs:
        - Info: Successful OTP resend
        - Warning: Invalid session or action
        - Error: Failed resend attempts and exceptions
    """
    from loguru import logger

    try:
        data = json.loads(request.body)  # Parse the JSON body
        username = request.session.get("username")

        if not username:
            logger.warning("Username not found in session")
            return JsonResponse(
                {"success": False, "message": "User session expired"}, status=400
            )

        user = User.objects.get(username=username)
        logger.info(f"User found: {username}")

        if "action" in data and data["action"] == "resend":
            logger.info("Resend OTP action received")
            # Logic to resend OTP, for example:
            success = send_otp_via_email(user)  # Replace with actual OTP logic

            if success:
                logger.info(f"OTP successfully sent to {username}")
                return JsonResponse({"success": True})
            else:
                logger.error(f"Failed to resend OTP to {username}")
                return JsonResponse(
                    {"success": False, "message": "Failed to resend OTP"}
                )
        else:
            logger.warning("Invalid action received")
            return JsonResponse(
                {"success": False, "message": "Invalid action"}, status=400
            )

    except json.JSONDecodeError:
        logger.error("Invalid JSON data received")
        return JsonResponse(
            {"success": False, "message": "Invalid JSON data"}, status=400
        )
    except User.DoesNotExist:
        logger.error(f"User with username {username} not found")
        return JsonResponse({"success": False, "message": "User not found"}, status=404)
    except Exception as e:
        logger.exception(f"Unexpected error in resend_otp: {str(e)}")
        return JsonResponse({"success": False, "message": str(e)}, status=500)


def contact_view(request):
    """
    Handle contact form submissions, save to database, and send email notifications.

    This view processes contact form submissions by:
    - Validating required form fields
    - Saving message to database
    - Formatting and sending notification email
    - Handling validation and email sending errors

    Args:
        request: HttpRequest object containing POST data with contact form fields

    Returns:
        JsonResponse: Success/failure status and message

    Status Codes:
        200: Success
        400: Missing fields or validation error
        500: Email sending error

    Logs:
        - Info: Successful form submissions
        - Warning: Invalid email addresses
        - Error: Failed email sending attempts
    """
    logger.debug("Processing contact form submission")

    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        subject = request.POST.get("subject")
        message = request.POST.get("message")

        logger.debug(
            f"Received form data - Name: {name}, Email: {email}, Subject: {subject}"
        )

        if not name or not email or not subject or not message:
            logger.warning("Missing required fields in contact form submission")
            return JsonResponse(
                {"status": "error", "message": "All fields are required."}, status=400
            )

        # Save message to database
        try:
            from .models import ContactMessage
            contact_msg = ContactMessage.objects.create(
                name=name,
                email=email,
                subject=subject,
                message=message,
                status='new'
            )
            logger.info(f"Saved contact message to database - ID: {contact_msg.id}")
        except Exception as e:
            logger.error(f"Error saving contact message to database: {str(e)}")

        # Create a well-formatted email body
        email_message = f"""
        You have received a new message from the contact form on your website:

        Name: {name}
        Email: {email}

        Subject: {subject}

        Message:
        {message}

        -----
        This email was automatically generated by your website.
        """

        try:
            send_mail(
                subject=f"[Website Contact Form] {subject}",  # Add a prefix for easy filtering
                message=email_message,
                from_email=settings.DEFAULT_FROM_EMAIL,  # Reply-To address (user's email)
                recipient_list=[os.getenv("EMAIL_HOST_USER")],  # Send to your Gmail address
            )
            logger.info(f"Successfully sent contact form email from {email}")
            return JsonResponse(
                {"status": "success", "message": "Your message has been sent!"}
            )
        except ValidationError as ve:
            logger.warning(f"Invalid email address: {email}")
            return JsonResponse({"status": "error", "message": str(ve)}, status=400)
        except Exception as e:
            logger.exception(f"Error sending contact form email: {str(e)}")
            return JsonResponse({"status": "error", "message": str(e)})


@login_required(login_url="/login/")
@admin_required
def admin_panel(request):
    """
    Render the admin panel with user management interface.

    This view handles displaying the admin panel by:
    - Retrieving the current admin user from session
    - Fetching all users from the database
    - Rendering the admin panel template with user data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered admin_panel.html template with user data

    Logs:
        - Info: When admin user accesses the panel
        - Warning: When username exists in session but user not found in DB
        - Debug: When fetching user list
    """
    logger.debug("Processing admin panel request")

    uname = request.session.get("username")
    # Initialize user and role variables in case user is not logged in
    user = None

    # Check if username exists in session and fetch user and role
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"Admin user {uname} accessed the admin panel")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None

    users = User.objects.all()
    logger.debug(f"Retrieved {len(users)} users from database")

    # Latest membership (etc.) payment vs AMC — keep separate so AMC seed does not hide membership row
    latest_by_user_id = {}
    amc_by_user_id = {}
    for payment in (
        PaymentRegistration.objects.filter(user__isnull=False)
        .order_by("user_id", "-created_at")
        .only("id", "user_id", "payment_status", "receipt", "amount", "notes", "updated_at", "created_at")
    ):
        uid = payment.user_id
        if _is_amc_payment_registration(payment):
            if uid not in amc_by_user_id:
                amc_by_user_id[uid] = payment
        else:
            if uid not in latest_by_user_id:
                latest_by_user_id[uid] = payment

    for u in users:
        u.latest_payment = latest_by_user_id.get(u.id)
        u.amc_payment = amc_by_user_id.get(u.id)

    # Calculate statistics
    total_users = User.objects.count()
    approved_count = User.objects.filter(status='approved').count()
    pending_count = User.objects.filter(status='pending').count()
    rejected_count = User.objects.filter(status='rejected').count()
    
    # Calculate category statistics (human-readable labels from model choices)
    from django.db.models import Count
    category_stats = User.objects.values('category').annotate(count=Count('id')).order_by('category')
    cat_labels = dict(_user_category_choice_tuples())
    by_category = [
        {
            "key": item["category"],
            "label": cat_labels.get(item["category"], item["category"]),
            "count": item["count"],
        }
        for item in category_stats
        if item["category"]
    ]

    statistics = {
        'total_users': total_users,
        'by_status': {
            'approved': approved_count,
            'pending': pending_count,
            'rejected': rejected_count,
        },
        'by_category': by_category,
    }
    
    logger.debug(f"Statistics calculated - Total: {total_users}, Approved: {approved_count}, Pending: {pending_count}, Rejected: {rejected_count}")

    data = {
        "user": user,
        "users": users,
        "statistics": statistics,
        "category_choices": _user_category_choice_tuples(),
    }
    logger.debug("Rendering admin panel")
    return render(request, "admin_panel.html", {"data": data})


@login_required(login_url="/login/")
@admin_required
def admin_bulk_users_template(request):
    """Download blank XLSX matching bulk user import column layout."""
    bio = build_bulk_user_template_bytes()
    resp = HttpResponse(
        bio.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="adams_bulk_users_template.xlsx"'
    return resp


@login_required(login_url="/login/")
@admin_required
@require_POST
def admin_bulk_users_upload(request):
    """Create users from uploaded XLSX (same columns as template)."""
    f = request.FILES.get("bulk_users_xlsx")
    if not f:
        messages.error(request, "Please choose an .xlsx file to upload.")
        return redirect("admin_panel")
    if not f.name.lower().endswith(".xlsx"):
        messages.error(request, "Only .xlsx files are accepted.")
        return redirect("admin_panel")
    try:
        created, errs, created_accounts = import_users_from_xlsx(f)
    except Exception as e:
        logger.exception("Bulk user import failed")
        messages.error(request, f"Could not read the spreadsheet: {e}")
        return redirect("admin_panel")
    if created_accounts:
        login_url = request.build_absolute_uri(reverse("login"))
        profile_url = request.build_absolute_uri(reverse("user_profile"))
        upload_url = request.build_absolute_uri(reverse("member_complete_uploads"))
        for acc in created_accounts:
            try:
                send_new_account_credentials_email(
                    username=acc["username"],
                    email=acc["email"],
                    password_plain=acc.get("plain_password") or "",
                    login_url=login_url,
                    profile_url=profile_url,
                    upload_url=upload_url,
                )
            except Exception:
                logger.exception(
                    "Welcome email failed for bulk-created user %s",
                    acc.get("email"),
                )
    if created:
        messages.success(request, f"Successfully created {created} user(s).")
    if errs:
        preview = " | ".join(errs[:20])
        if len(errs) > 20:
            preview += f" …and {len(errs) - 20} more (see logs)."
        messages.warning(request, f"Some rows were skipped ({len(errs)}): {preview}")
    if not created and not errs:
        messages.info(request, "No data rows found below the header.")
    return redirect("admin_panel")


@login_required(login_url="/login/")
def user_profile(request):
    """
    Render the user profile page with user data.

    This view handles displaying the user profile page by:
    - Retrieving the current user from session if logged in
    - Rendering the user profile template with user data
    - Checking for pending category change requests

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered user_profile.html template with user data

    Logs:
        - Info: When user accesses their profile page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing user profile page request")

    uname = request.session.get("username")
    # Initialize user and role variables in case user is not logged in
    user = None
    pending_category_request = None
    pending_category_labels = None
    latest_payment = None
    amc_payment = None

    # Check if username exists in session and fetch user and role
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed their profile page")
            
            # Check for pending category change request
            pending_category_request = CategoryChangeRequest.objects.filter(
                user=user, 
                request_status='pending'
            ).first()

            if pending_category_request:
                cmap = dict(_user_category_choice_tuples())
                pending_category_labels = {
                    "current": cmap.get(
                        pending_category_request.current_category,
                        pending_category_request.current_category,
                    ),
                    "requested": cmap.get(
                        pending_category_request.requested_category,
                        pending_category_request.requested_category,
                    ),
                }

            _pay_qs = PaymentRegistration.objects.filter(user=user).order_by("-created_at")
            for p in _pay_qs:
                if _is_amc_payment_registration(p):
                    if amc_payment is None:
                        amc_payment = p
                else:
                    if latest_payment is None:
                        latest_payment = p
                if latest_payment is not None and amc_payment is not None:
                    break
            
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    upload_gaps = {}
    needs_any_upload = False
    amc_pay_details = None
    show_pay_amc_button = False
    if user:
        upload_gaps = member_profile_file_gaps(user)
        needs_any_upload = any(upload_gaps.values())
        amc_pay_details = _get_membership_details("amc")
        # Offer checkout whenever AMC is not recorded as paid (missing, pending, or failed).
        show_pay_amc_button = bool(
            amc_pay_details
            and (amc_payment is None or amc_payment.payment_status != "paid")
        )

    data = {
        "user": user,
        "pending_category_request": pending_category_request,
        "pending_category_labels": pending_category_labels,
        "latest_payment": latest_payment,
        "amc_payment": amc_payment,
        "amc_pay_details": amc_pay_details,
        "show_pay_amc_button": show_pay_amc_button,
        "category_choices": _user_category_choice_tuples(),
        "upload_gaps": upload_gaps,
        "needs_any_upload": needs_any_upload,
    }
    logger.debug("Rendering user profile page")
    return render(request, "user_profile.html", {"data": data})


def _save_user_media_file_to_disk(user, field_name, uploaded_file):
    """Write an uploaded file under MEDIA_ROOT and return the stored path (media/...)."""
    folder_path = {
        "photo": f"users/{user.username}/photo/",
        "state_nmc": f"users/{user.username}/state_nmc/",
        "passport": f"users/{user.username}/passport/",
        "medical_qualification": f"users/{user.username}/medical_qualification/",
        "payment_transaction_proof": f"users/{user.username}/payment_transaction_proof/",
    }[field_name]
    file_name = f"{folder_path}{uploaded_file.name}"
    media_path = os.path.join(settings.MEDIA_ROOT, file_name)
    os.makedirs(os.path.dirname(media_path), exist_ok=True)
    uploaded_file.seek(0)
    with open(media_path, "wb+") as destination:
        for chunk in uploaded_file.chunks():
            destination.write(chunk)
    return f"media/{file_name}"


def member_complete_uploads(request):
    """
    Let the logged-in member upload a profile photo and/or required documents
    only for fields that are still missing or bulk placeholders.
    """
    uname = request.session.get("username")
    if not uname:
        messages.error(request, "Please log in to upload files.")
        return redirect("login")

    try:
        user = User.objects.get(username=uname)
    except User.DoesNotExist:
        messages.error(request, "Account not found.")
        return redirect("login")

    field_labels = {
        "photo": "Passport-size photo (used on your membership ID card)",
        "passport": "Passport (front and back)",
        "medical_qualification": "Medical qualification / student ID (PDF or image)",
        "payment_transaction_proof": "Payment transaction proof / screenshot",
        "state_nmc": "State / NMC registration certificate (if applicable)",
    }

    gaps = member_profile_file_gaps(user)

    if request.method == "POST":
        gaps_now = member_profile_file_gaps(user)
        saved = 0
        for field_name in request.FILES:
            if field_name not in MEMBER_PROFILE_FILE_FIELDS:
                continue
            if not gaps_now.get(field_name):
                messages.warning(
                    request,
                    f"{field_name.replace('_', ' ').title()} is already on file and was not changed.",
                )
                continue
            uploaded = request.FILES[field_name]
            try:
                rel_path = _save_user_media_file_to_disk(user, field_name, uploaded)
                setattr(user, field_name, rel_path)
                user.save(update_fields=[field_name])
                saved += 1
                logger.info(
                    "Member %s uploaded missing file field=%s path=%s",
                    user.username,
                    field_name,
                    rel_path,
                )
            except Exception as e:
                logger.exception("Member upload failed for %s: %s", field_name, e)
                messages.error(
                    request,
                    f"Could not save {field_name.replace('_', ' ')}: {e}",
                )
        if saved:
            messages.success(
                request,
                f"Saved {saved} file(s). Thank you for completing your profile.",
            )
            return redirect("user_profile")

        messages.error(
            request,
            "No new files were saved. Choose a file for at least one item that still shows as required.",
        )
        gaps = member_profile_file_gaps(user)

    if not any(gaps.values()):
        messages.info(request, "Your profile photo and documents are already on file.")
        return redirect("user_profile")

    missing_fields = [
        (field_name, field_labels[field_name])
        for field_name in MEMBER_PROFILE_FILE_FIELDS
        if gaps.get(field_name)
    ]
    return render(
        request,
        "member_complete_uploads.html",
        {
            "user": user,
            "missing_fields": missing_fields,
        },
    )


def member_pay_amc(request):
    """Start Razorpay checkout for AMC (₹500) for the logged-in portal user."""
    user = _portal_session_user(request)
    if not user:
        messages.error(request, "Please log in to pay AMC.")
        return redirect("login")
    phone = (user.mobile_number or "").strip()
    if not phone:
        messages.error(request, "Add a mobile number on your profile before paying AMC.")
        return redirect("user_profile")
    details = _get_membership_details("amc")
    if not details:
        messages.error(request, "AMC payment is not available right now.")
        return redirect("user_profile")
    request.session["payment_prefill"] = {
        "name": f"{user.first_name} {user.last_name}".strip() or user.username,
        "email": user.email,
        "phone": phone,
        "membership_type": "amc",
        "membership_label": details["label"],
        "amount": details["amount"],
    }
    return redirect("registration_payment_redirect")


def member_payment_history(request):
    """All Razorpay / payment rows for the current portal user."""
    user = _portal_session_user(request)
    if not user:
        messages.error(request, "Please log in to view payment history.")
        return redirect("login")
    q = Q(user=user)
    if (user.email or "").strip():
        q |= Q(user__isnull=True, email__iexact=user.email.strip())
    payments = list(
        PaymentRegistration.objects.filter(q).order_by("-created_at")
    )
    return render(
        request,
        "member_payment_history.html",
        {"user": user, "payments": payments},
    )


def member_payment_receipt(request, payment_id):
    """HTML receipt for one payment (view or download). Only the owning member may access."""
    user = _portal_session_user(request)
    if not user:
        messages.error(request, "Please log in to download your receipt.")
        return redirect("login")
    payment = get_object_or_404(PaymentRegistration, pk=payment_id)
    if not _payment_registration_visible_to_user(payment, user):
        raise Http404("Receipt not found.")

    kind_label = _payment_receipt_kind_label(payment)

    as_download = request.GET.get("download") in ("1", "true", "yes")
    response = render(
        request,
        "payment_receipt.html",
        {
            "payment": payment,
            "member": user,
            "kind_label": kind_label,
            "as_download": as_download,
        },
    )
    if as_download:
        stem = slugify(payment.receipt or f"payment-{payment.pk}")[:80] or "receipt"
        response["Content-Disposition"] = f'attachment; filename="adams-receipt-{stem}.html"'
    return response


@login_required(login_url="/login/")
@admin_required
def admin_member_payment_history(request, user_id):
    """Admin: full payment history for one member."""
    member = get_object_or_404(User, pk=user_id)
    q = Q(user=member)
    if (member.email or "").strip():
        q |= Q(user__isnull=True, email__iexact=member.email.strip())
    payments = list(
        PaymentRegistration.objects.filter(q).order_by("-created_at")
    )
    return render(
        request,
        "admin_member_payment_history.html",
        {"member": member, "payments": payments},
    )


@login_required(login_url="/login/")
@admin_required
def admin_member_payment_receipt(request, user_id, payment_id):
    """Admin: download/view receipt for one of a member's payments (same HTML as member receipt)."""
    member = get_object_or_404(User, pk=user_id)
    payment = get_object_or_404(PaymentRegistration, pk=payment_id)
    if not _payment_registration_belongs_to_member(payment, member):
        raise Http404("Receipt not found.")

    kind_label = _payment_receipt_kind_label(payment)
    as_download = request.GET.get("download") in ("1", "true", "yes")
    response = render(
        request,
        "payment_receipt.html",
        {
            "payment": payment,
            "member": member,
            "kind_label": kind_label,
            "as_download": as_download,
        },
    )
    if as_download:
        stem = slugify(f"{member.username}-{payment.receipt or payment.pk}")[:80] or "receipt"
        response["Content-Disposition"] = f'attachment; filename="adams-receipt-{stem}.html"'
    return response


def _id_card_render_context(user, *, id_card_png_url=None):
    """Shared context for member_id_card.html (photo, name, category, membership id, QR data URI)."""
    if id_card_png_url is None:
        id_card_png_url = reverse("member_id_card_png")
    membership_id = user.display_membership_id
    name_parts = [
        (user.first_name or "").strip(),
        (user.middle_name or "").strip(),
        (user.last_name or "").strip(),
    ]
    member_display_name = " ".join(p for p in name_parts if p) or user.username
    qr_data_uri = None
    try:
        qr_buf = generate_qr_code_live(user)
        qr_data_uri = "data:image/png;base64," + base64.b64encode(qr_buf.getvalue()).decode(
            "ascii"
        )
    except Exception as e:
        logger.warning(f"Could not embed QR on ID card for {user.username}: {e}")
    return {
        "member": user,
        "membership_id": membership_id,
        "member_display_name": member_display_name,
        "qr_data_uri": qr_data_uri,
        # Bulk (and other) accounts may still have a placeholder path in `photo`; avoid a broken <img>.
        "has_real_photo": not stored_file_is_missing(user.photo),
        "id_card_png_url": id_card_png_url,
    }


def member_id_card(request):
    """
    Printable / downloadable membership ID card for approved members (photo, name, heading, category, ID, QR).
    Uses portal session (same as home page), not only Django's login_required user.
    """
    user = _portal_session_user(request)
    if not user:
        messages.error(request, "Please log in to view your ID card.")
        return redirect("login")
    if user.status != "approved":
        messages.warning(
            request,
            "Your membership ID card will be available after your application is approved.",
        )
        return redirect("user_profile")
    return render(request, "member_id_card.html", _id_card_render_context(user))


def member_id_card_png(request):
    """Download membership ID card as PNG (server-rendered; works for bulk-import placeholders)."""
    user = _portal_session_user(request)
    if not user:
        return HttpResponse(status=401, content="Please log in (session expired).")
    if user.status != "approved":
        return HttpResponse(
            status=403,
            content="Membership ID card download is available after your application is approved.",
        )
    try:
        data = build_id_card_png_bytes(user)
    except Exception as e:
        logger.exception("member_id_card_png failed for %s: %s", uname, e)
        return HttpResponse(status=500, content="Could not generate image")
    fn = id_card_png_filename(user)
    resp = HttpResponse(data, content_type="image/png")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


@login_required(login_url="/login/")
@admin_required
def admin_member_id_card_png(request, user_id):
    """Admin: download member ID card PNG (server-rendered)."""
    member = get_object_or_404(User, pk=user_id)
    try:
        data = build_id_card_png_bytes(member)
    except Exception as e:
        logger.exception("admin_member_id_card_png failed for user_id=%s: %s", user_id, e)
        return HttpResponse(status=500, content="Could not generate image")
    fn = id_card_png_filename(member)
    resp = HttpResponse(data, content_type="image/png")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    return resp


@login_required(login_url="/login/")
@admin_required
def admin_member_id_card(request, user_id):
    """Admin: open printable/downloadable ID card for any member (same layout as self-service ID card)."""
    member = get_object_or_404(User, pk=user_id)
    logger.info(
        "Admin %s opened member ID card for user_id=%s username=%s",
        getattr(request.user, "username", ""),
        user_id,
        member.username,
    )
    ctx = _id_card_render_context(
        member,
        id_card_png_url=reverse("admin_member_id_card_png", args=[member.pk]),
    )
    ctx["id_card_back_url"] = reverse("admin_panel")
    ctx["id_card_back_label"] = "Back to admin panel"
    return render(request, "member_id_card.html", ctx)


@login_required(login_url="/login/")
def retry_payment(request):
    uname = request.session.get("username")
    if not uname:
        messages.error(request, "Please login to retry payment.")
        return redirect("login")

    try:
        user = User.objects.get(username=uname)
    except User.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect("login")

    latest_payment = (
        PaymentRegistration.objects.filter(user=user)
        .order_by("-created_at")
    )
    latest_payment = next(
        (p for p in latest_payment if not _is_amc_payment_registration(p)),
        None,
    )

    membership_type = None
    membership_label = ""
    amount = MEMBERSHIP_TYPES["adhoc"]["amount"]

    if latest_payment and isinstance(latest_payment.notes, dict):
        membership_type = (latest_payment.notes.get("membership_type") or "").strip().lower() or None

    membership_details = _get_membership_details(membership_type) if membership_type else None
    if membership_details:
        membership_label = membership_details["label"]
        amount = membership_details["amount"]
        membership_type = membership_type or ""
    else:
        membership_type = "adhoc"
        membership_label = MEMBERSHIP_TYPES["adhoc"]["label"]

    request.session["payment_prefill"] = {
        "name": f"{user.first_name} {user.last_name}".strip() or user.username,
        "email": user.email,
        "phone": user.mobile_number,
        "membership_type": membership_type,
        "membership_label": membership_label,
        "amount": amount,
    }
    messages.info(request, "Retrying payment...")
    return redirect("registration_payment_redirect")


@login_required(login_url="/login/")
def reset_password(request):
    """
    Handle password reset requests for logged in users.

    This view processes password reset requests by:
    - Validating old password matches current password
    - Validating new password and confirmation match
    - Updating the user's password if validation passes

    Args:
        request: HttpRequest object containing POST data with passwords

    Returns:
        HttpResponseRedirect: Redirect to user profile with success/error message
        HttpResponse: Rendered user profile template for GET requests

    Logs:
        - Info: Successful password resets
        - Warning: Validation failures
        - Error: Database errors
    """
    logger.debug("Processing password reset request")

    if request.method == "POST":
        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        if not old_password or not new_password or not confirm_password:
            logger.warning("Missing required password fields")
            messages.error(request, "All fields are required.")
            return redirect("user_profile")

        if new_password != confirm_password:
            logger.warning("New password and confirmation do not match")
            messages.error(request, "New password and confirm password do not match.")
            return redirect("user_profile")

        try:
            user = User.objects.get(username=request.session.get("username"))

            if not user.check_password(old_password):
                logger.warning(
                    f"Invalid old password provided for user: {user.username}"
                )
                messages.error(request, "Old password is incorrect.")
                return redirect("user_profile")

            user.set_password(new_password)
            user.save()
            logger.info(f"Password successfully updated for user: {user.username}")
            messages.success(request, "Password updated successfully.")
            return redirect("user_profile")

        except User.DoesNotExist:
            logger.error(
                f"User not found for username: {request.session.get('username')}"
            )
            messages.error(request, "User not found.")
            return redirect("user_profile")
        except Exception as e:
            logger.exception(f"Error updating password: {str(e)}")
            messages.error(request, "An error occurred. Please try again.")
            return redirect("user_profile")

    logger.debug("Rendering password reset form")
    return render(request, "reset_password.html")


@login_required(login_url="/login/")
def get_user_details(request, userId):
    """
    Get detailed information for a specific user.

    This view retrieves and returns detailed user information by:
    - Fetching user object by ID
    - Converting user data to JSON format
    - Handling user not found errors

    Args:
        request: HttpRequest object
        userId (int): ID of the user to fetch details for

    Returns:
        JsonResponse: User details in JSON format

    Raises:
        Http404: If user with given ID is not found

    Logs:
        - Info: When user details are successfully retrieved
        - Error: When user is not found
    """
    logger.debug(f"Fetching user details for ID: {userId}")

    try:
        user = get_object_or_404(User, id=userId)
        logger.info(f"Retrieved user details for ID {userId}")

        data = {
            "success": True,
            "user": {
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "gender": user.gender,
                "mobile_number": user.mobile_number,
                "district": user.district,
                "role": user.role,
                "whatsapp_number": user.whatsapp_number,
                "address_communication": user.address_communication,
                "address_permanent": user.address_permanent,
                "father_spouse_details": user.father_spouse_details,
                "blood_group": user.blood_group,
                "educational_status": user.educational_status,
                "category": user.category,
                "category_display": user.get_category_display(),
                "membership_id": user.display_membership_id,
                "university_name": user.university_name,
                "country_university": user.country_university,
                "year_of_joining": user.year_of_joining,
                "year_of_completion": user.year_of_completion,
                "date_time_of_payment": user.date_time_of_payment,
                "willing_to_be_donor": user.willing_to_be_donor,
                # "agreement": user.agreement,
                "mid": user.mid,
                "status": user.status,
                "admin_remarks": user.admin_remarks,
            },
        }
        logger.debug(f"Returning user details for ID {userId}")
        return JsonResponse(data)

    except User.DoesNotExist:
        logger.error(f"User not found with ID: {userId}")
        raise Http404("User not found")


def fetch_image_from_minio(bucket_name, object_name):
    """
    Fetch an image file from local storage (backward compatibility function).
    
    This function now retrieves files from local storage instead of MinIO.
    Kept for backward compatibility with existing code.

    Args:
        bucket_name (str): Ignored (kept for compatibility)
        object_name (str): Path/name of the file

    Returns:
        bytes: Binary content of the file if successful, None if error

    Logs:
        - Info: When file is successfully retrieved
        - Error: When file fetch fails
    """
    logger.debug(f"Attempting to fetch file {object_name} from local storage")
    try:
        # If path doesn't start with "media/", add it
        if not object_name.startswith("media/"):
            object_name = f"media/{object_name}"
        
        full_path = os.path.join(settings.BASE_DIR, object_name)
        full_path = os.path.normpath(full_path)
        
        # Security check
        media_root = os.path.normpath(settings.MEDIA_ROOT)
        if not full_path.startswith(media_root):
            logger.error(f"Security violation: File path outside MEDIA_ROOT: {object_name}")
            return None
        
        if not os.path.exists(full_path):
            logger.error(f"File not found: {full_path}")
            return None
        
        with open(full_path, 'rb') as f:
            file_data = f.read()
        
        logger.info(f"Successfully retrieved file {object_name} from local storage")
        return file_data
    except Exception as e:
        logger.error(f"Error fetching file {object_name} from local storage: {str(e)}")
        return None


_SERVE_USER_FILE_FIELDS = frozenset(
    {
        "photo",
        "state_nmc",
        "passport",
        "medical_qualification",
        "payment_transaction_proof",
    }
)


def serve_user_file(request, user_id, file_field):
    """
    Serve a user's file (image or PDF) from MinIO storage or local filesystem.

    Auth: portal session user (same as home page) or Django authenticated user;
    only that user may fetch their files unless the viewer is an admin.

    This view retrieves and serves files associated with a user by:
    - Fetching the user object and requested file field
    - Checking if file is stored locally (media/) or in MinIO
    - Getting the file from appropriate storage location
    - Setting appropriate content type headers
    - Returning file data as HTTP response

    Args:
        request: HttpRequest object
        user_id (int): ID of the user whose file is requested
        file_field (str): Name of the file field to retrieve

    Returns:
        HttpResponse: File data with appropriate content type if successful
        JsonResponse: Error message if file not found
        HttpResponse: Error status if file fetch fails

    Raises:
        Http404: If user with given ID is not found

    Logs:
        - Info: When file is successfully served
        - Warning: When file field is empty
        - Error: When file fetch fails or type is unsupported
    """
    logger.debug(
        f"Processing file serve request for user ID {user_id}, field {file_field}"
    )

    viewer = _portal_session_user(request)
    if not viewer:
        return HttpResponse(status=401, content="Unauthorized")
    if file_field not in _SERVE_USER_FILE_FIELDS:
        return HttpResponse(status=400, content="Invalid file field")
    if viewer.pk != user_id and getattr(viewer, "role", None) != "admin":
        return HttpResponse(status=403, content="Forbidden")

    user = get_object_or_404(User, id=user_id)
    file_path = getattr(user, file_field, None)

    # Handle empty/None file paths (especially for optional fields like state_nmc)
    if not file_path or file_path == "" or str(file_path).strip() == "" or str(file_path) == "None":
        logger.debug(f"File field {file_field} is empty for user {user_id} (optional field may not be provided)")
        return JsonResponse({"error": "File not provided"}, status=404)

    # All files are stored locally now - serve from filesystem
    file_path_str = str(file_path)
    
    # If path doesn't start with "media/", add it for backward compatibility
    if not file_path_str.startswith("media/"):
        file_path_str = f"media/{file_path_str}"
    
    logger.debug(f"Serving local file: {file_path_str}")
    try:
        # Construct full file path
        # BASE_DIR is a Path object, convert to string
        base_dir = str(settings.BASE_DIR)
        full_path = os.path.join(base_dir, file_path_str)
        
        # Normalize path separators for Windows
        full_path = os.path.normpath(full_path)
        
        # Security check: ensure file is within MEDIA_ROOT
        media_root = os.path.normpath(settings.MEDIA_ROOT)
        if not full_path.startswith(media_root):
            logger.error(f"Security violation: File path outside MEDIA_ROOT: {file_path_str}")
            return HttpResponse(status=403, content="Access denied")
        
        if not os.path.exists(full_path):
            logger.error(f"Local file not found: {full_path}")
            return JsonResponse({"error": "File not found"}, status=404)
        
        # Read file
        with open(full_path, 'rb') as f:
            file_data = f.read()
        
        logger.info(f"Successfully read local file: {full_path} (size: {len(file_data)} bytes)")
        
        # Determine content type
        if file_path_str.endswith(".pdf"):
            content_type = "application/pdf"
        elif file_path_str.endswith((".jpg", ".jpeg")):
            content_type = "image/jpeg"
        elif file_path_str.endswith(".png"):
            content_type = "image/png"
        else:
            content_type = "application/octet-stream"
        
        logger.info(f"Successfully serving local {content_type} file for user {user_id}")
        return HttpResponse(file_data, content_type=content_type)
        
    except Exception as e:
        logger.exception(f"Error serving local file {file_path}: {e}")
        return HttpResponse(status=500, content=f"Error serving file: {str(e)}")


@csrf_exempt
@login_required(login_url="/login/")
@admin_required
def update_user_status(request, user_id):
    """
    Update a user's status and admin remarks.

    This view handles updating user status and admin remarks by:
    - Validating the request contains proper JSON data
    - Updating the user's status and admin remarks in the database
    - Handling validation and database errors

    Args:
        request: HttpRequest object containing POST data with status and remarks
        user_id (int): ID of the user to update

    Returns:
        JsonResponse: Success/failure status and message

    Status Codes:
        200: Success
        400: Invalid request data
        404: User not found
        500: Database error

    Logs:
        - Info: Successful status updates
        - Warning: Invalid request methods
        - Error: Failed updates and exceptions
    """
    logger.debug(f"Processing status update request for user ID: {user_id}")

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            status = data.get("status")
            admin_remarks = data.get("admin_remarks", "")

            logger.debug(
                f"Received status update - Status: {status}, Remarks: {admin_remarks}"
            )

            # Update the user object
            user = User.objects.get(id=user_id)
            user.status = status
            user.admin_remarks = admin_remarks  # Update admin remarks
            user.save()

            if status == "rejected":
                # Delete the user's photo from MinIO
                send_email_with_attachments(
                    subject="Membership Request Update",
                    custom_message=f"""Thank you for your interest in joining the Association of Doctors and Medical Students (ADAMS).  
                                        <br><br>
                                        After a thorough review of your application, we regret to inform you that your membership request has not been approved at this time.  
                                        <br><br>
                                        <strong>Admin Remarks:</strong> <em>{admin_remarks}</em>  
                                        <br><br>
                                        We genuinely appreciate the effort and time you invested in your application.  
                                        <br><br>
                                        Should you have any questions or need further clarification, feel free to reach out to us at <a href="https://adams.org.in/contact" style="color: #3498db; text-decoration: none;">Contact Us</a>.  
                                        <br><br>
                                        Warm regards,  
                                        <br>
                                        <strong>Membership Team</strong>  
                                        <br>
                                        <strong>Association of Doctors and Medical Students (ADAMS)</strong>
                                        """,
                    template_type="custom",
                    recipient_list=[user.email],
                )
            elif status == "approved":
                send_email_with_attachments(
                    subject="Membership Request Update",
                    custom_message="""Thank you for your interest in joining the Association of Doctors and Medical Students (ADAMS).
                    We are pleased to inform you that your membership request has been approved.
                    Should you have any questions or need further clarification, please don't hesitate to reach out.""",
                    template_type="custom",
                    recipient_list=[user.email],
                )

            logger.info(
                f"Successfully updated status to {status} for user ID {user_id}"
            )
            return JsonResponse(
                {
                    "success": True,
                    "message": "User status and remarks updated successfully.",
                }
            )
        except User.DoesNotExist:
            logger.error(f"User not found with ID: {user_id}")
            return JsonResponse({"success": False, "message": "User not found."})
        except Exception as e:
            logger.exception(f"Error updating user {user_id} status: {str(e)}")
            return JsonResponse({"success": False, "message": str(e)})
    logger.warning(f"Invalid request method for user {user_id} status update")
    return JsonResponse({"success": False, "message": "Invalid request method."})


def render_about_us(request):
    """
    Render the about us page with user data.

    This view handles displaying the about us page by:
    - Retrieving the current user from session if logged in
    - Rendering the about us template with user data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered aboutus.html template with user data

    Logs:
        - Info: When user accesses about us page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing about us page request")

    uname = request.session.get("username")
    # Initialize user and role variables in case user is not logged in
    user = None

    # Check if username exists in session and fetch user and role
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the about us page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    data = {"user": user}
    logger.debug("Rendering about us page")
    return render(request, "aboutus.html", {"data": data})


def render_contact_us(request):
    """
    Render the contact us page with user data.

    This view handles displaying the contact us page by:
    - Retrieving the current user from session if logged in
    - Rendering the contact us template with user data

    Args:
        request: HttpRequest object containing session and other metadata

    Returns:
        HttpResponse: Rendered contact_us.html template with user data

    Logs:
        - Info: When user accesses contact us page
        - Warning: When username exists in session but user not found in DB
        - Debug: When no username found in session
    """
    logger.debug("Processing contact us page request")

    uname = request.session.get("username")
    # Initialize user and role variables in case user is not logged in
    user = None

    # Check if username exists in session and fetch user and role
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the contact us page")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")

    data = {"user": user}
    logger.debug("Rendering contact us page")
    return render(request, "contact_us.html", {"data": data})


@csrf_exempt
@login_required(login_url="/login/")
def request_category_change(request):
    """
    Handle member's request to change their category.

    This view allows members to request a category change by:
    - Validating the request contains proper JSON data
    - Creating a new category change request
    - Sending notification email to admins

    Args:
        request: HttpRequest object containing POST data with new category

    Returns:
        JsonResponse: Success/failure status and message

    Status Codes:
        200: Success
        400: Invalid request data or existing pending request
        404: User not found
        500: Server error

    Logs:
        - Info: Successful category change requests
        - Warning: Duplicate pending requests
        - Error: Failed requests and exceptions
    """
    logger.debug("Processing category change request")
    
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_category = (data.get("new_category") or "").strip()
            allowed_cat = _user_category_values()
            if new_category not in allowed_cat:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Invalid membership category selected.",
                    },
                    status=400,
                )

            uname = request.session.get("username")
            user = User.objects.get(username=uname)
            if new_category == user.category:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Please choose a category different from your current one.",
                    },
                    status=400,
                )

            # Check if there's already a pending request
            existing_request = CategoryChangeRequest.objects.filter(
                user=user,
                request_status='pending'
            ).first()
            
            if existing_request:
                logger.warning(f"User {user.username} has a pending category change request")
                return JsonResponse({
                    "success": False,
                    "message": "You already have a pending category change request."
                }, status=400)
            
            # Create new category change request
            category_request = CategoryChangeRequest.objects.create(
                user=user,
                current_category=user.category,
                requested_category=new_category
            )
            
            logger.info(f"Category change request created for user {user.username}: {user.category} -> {new_category}")
            messages.success(request, "Category change request submitted successfully. Awaiting admin approval.")
            
            return JsonResponse({
                "success": True,
                "message": "Category change request submitted successfully."
            })
            
        except User.DoesNotExist:
            logger.error("User not found in session")
            return JsonResponse({
                "success": False,
                "message": "User not found."
            }, status=404)
        except Exception as e:
            logger.exception(f"Error creating category change request: {str(e)}")
            return JsonResponse({
                "success": False,
                "message": str(e)
            }, status=500)
    
    return JsonResponse({
        "success": False,
        "message": "Invalid request method."
    }, status=400)


@login_required(login_url="/login/")
@admin_required
def get_category_change_requests(request):
    """
    Get all pending category change requests for admin review.

    This view retrieves all pending category change requests by:
    - Fetching all requests with 'pending' status
    - Formatting the data for JSON response

    Args:
        request: HttpRequest object

    Returns:
        JsonResponse: List of pending category change requests

    Logs:
        - Info: Successful request retrieval
        - Error: Failed retrieval attempts
    """
    logger.debug("Fetching category change requests for admin")
    
    try:
        requests_list = CategoryChangeRequest.objects.filter(request_status='pending')
        
        data = []
        for req in requests_list:
            data.append({
                'id': req.id,
                'user_id': req.user.id,
                'username': req.user.username,
                'first_name': req.user.first_name,
                'last_name': req.user.last_name,
                'current_category': req.current_category,
                'requested_category': req.requested_category,
                'request_date': req.request_date.strftime("%Y-%m-%d %H:%M:%S"),
            })
        
        logger.info(f"Retrieved {len(data)} pending category change requests")
        return JsonResponse(
            {
                "success": True,
                "requests": data,
                "category_label_map": dict(_user_category_choice_tuples()),
            }
        )
        
    except Exception as e:
        logger.exception(f"Error fetching category change requests: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url="/login/")
@admin_required
def approve_reject_category_change(request, request_id):
    """
    Approve or reject a category change request.

    This view handles admin's decision on category change requests by:
    - Validating the decision (approve/reject)
    - Updating the user's category if approved
    - Sending notification email to the user
    - Recording admin remarks

    Args:
        request: HttpRequest object containing POST data with decision and remarks
        request_id: ID of the category change request

    Returns:
        JsonResponse: Success/failure status and message

    Status Codes:
        200: Success
        400: Invalid request data
        404: Request not found
        500: Server error

    Logs:
        - Info: Successful approvals/rejections
        - Warning: Invalid decisions
        - Error: Failed updates and exceptions
    """
    logger.debug(f"Processing category change decision for request ID: {request_id}")
    
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            decision = data.get("decision")  # 'approved' or 'rejected'
            admin_remarks = data.get("admin_remarks", "")
            
            category_request = CategoryChangeRequest.objects.get(id=request_id)
            user = category_request.user

            if decision == "approved" and category_request.requested_category not in _user_category_values():
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Requested category is not a valid membership category.",
                    },
                    status=400,
                )

            category_request.request_status = decision
            category_request.admin_decision_date = timezone.now()
            category_request.admin_remarks = admin_remarks
            category_request.save()
            
            if decision == 'approved':
                req_cat = category_request.requested_category
                # Update user's category
                old_category = user.category
                user.category = req_cat
                user.save()

                old_lbl = _category_display_name(old_category)
                new_lbl = _category_display_name(user.category)
                logger.info(f"Category change approved for user {user.username}: {old_category} -> {user.category}")

                # Send approval email
                send_email_with_attachments(
                    subject="Category Change Request Approved",
                    custom_message=f"""Your category change request has been approved!
                    <br><br>
                    Your category has been successfully updated from <strong>{old_lbl}</strong> to <strong>{new_lbl}</strong>.
                    <br><br>
                    {f'<strong>Admin Remarks:</strong> <em>{admin_remarks}</em><br><br>' if admin_remarks else ''}
                    Thank you for keeping your profile up to date.
                    <br><br>
                    Warm regards,<br>
                    <strong>ADAMS Team</strong>
                    """,
                    template_type="custom",
                    recipient_list=[user.email],
                )
                
                messages.success(request, f"Category change approved for {user.username}")
                
            elif decision == 'rejected':
                logger.info(f"Category change rejected for user {user.username}")
                cur_lbl = _category_display_name(category_request.current_category)
                req_lbl = _category_display_name(category_request.requested_category)

                # Send rejection email
                send_email_with_attachments(
                    subject="Category Change Request Update",
                    custom_message=f"""Your category change request has been reviewed.
                    <br><br>
                    Unfortunately, your request to change category from <strong>{cur_lbl}</strong> to <strong>{req_lbl}</strong> has not been approved.
                    <br><br>
                    <strong>Admin Remarks:</strong> <em>{admin_remarks if admin_remarks else 'No remarks provided'}</em>
                    <br><br>
                    If you have any questions, please contact us.
                    <br><br>
                    Warm regards,<br>
                    <strong>ADAMS Team</strong>
                    """,
                    template_type="custom",
                    recipient_list=[user.email],
                )
                
                messages.success(request, f"Category change rejected for {user.username}")
            
            return JsonResponse({
                "success": True,
                "message": f"Category change request {decision} successfully."
            })
            
        except CategoryChangeRequest.DoesNotExist:
            logger.error(f"Category change request not found: {request_id}")
            return JsonResponse({
                "success": False,
                "message": "Request not found."
            }, status=404)
        except Exception as e:
            logger.exception(f"Error processing category change decision: {str(e)}")
            return JsonResponse({
                "success": False,
                "message": str(e)
            }, status=500)
    
    return JsonResponse({
        "success": False,
        "message": "Invalid request method."
    }, status=400)


@login_required(login_url="/login/")
@admin_required
def get_category_members(request):
    """
    Get count and list of members by category.

    Args:
        request: HttpRequest object with category query parameter

    Returns:
        JsonResponse: Count and list of members in the category

    Logs:
        - Info: Successful member retrieval
        - Error: Failed retrieval attempts
    """
    logger.debug("Fetching category members")
    
    try:
        category = request.GET.get('category', '')
        allowed_cat = _user_category_values()

        if not category:
            return JsonResponse({
                'success': False,
                'message': 'Category parameter is required'
            }, status=400)

        if category != 'All' and category not in allowed_cat:
            return JsonResponse(
                {
                    'success': False,
                    'message': 'Invalid category. Use "All" or a value from the category list.',
                },
                status=400,
            )

        # Handle "All" categories
        if category == 'All':
            users = User.objects.all()
        else:
            users = User.objects.filter(category=category)
        
        count = users.count()
        
        logger.info(f"Retrieved {count} members for category: {category}")
        
        return JsonResponse({
            'success': True,
            'count': count,
            'category': category
        })
        
    except Exception as e:
        logger.exception(f"Error fetching category members: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required(login_url="/login/")
@admin_required
def export_category_members(request):
    """
    Export members of a specific category to Excel file.

    Args:
        request: HttpRequest object with category query parameter

    Returns:
        HttpResponse: Excel file download

    Logs:
        - Info: Successful export
        - Error: Export failures
    """
    logger.debug("Exporting category members to Excel")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        category = request.GET.get('category', '')
        allowed_cat = _user_category_values()

        if not category:
            return JsonResponse({
                'success': False,
                'message': 'Category parameter is required'
            }, status=400)

        if category != 'All' and category not in allowed_cat:
            return JsonResponse(
                {
                    'success': False,
                    'message': 'Invalid category. Use "All" or a value from the category list.',
                },
                status=400,
            )

        # Handle "All" categories
        if category == 'All':
            users = User.objects.all().order_by('category', 'first_name', 'last_name')
            sheet_title = "All Members"
        else:
            users = User.objects.filter(category=category).order_by('first_name', 'last_name')
            sheet_title = f"{_category_display_name(category)[:20]} Members"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]  # Excel sheet name limit is 31 characters
        
        # Add headers
        headers = [
            'Username', 'First Name', 'Last Name', 'Email', 'Gender',
            'Mobile Number', 'WhatsApp Number', 'District', 'Blood Group',
            'Educational Status', 'Category', 'University Name', 'Country',
            'Year of Joining', 'Year of Completion', 'Status',
            'ADAMS membership no.',
        ]
        
        # Style header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.username)
            ws.cell(row=row_num, column=2, value=user.first_name)
            ws.cell(row=row_num, column=3, value=user.last_name)
            ws.cell(row=row_num, column=4, value=user.email)
            ws.cell(row=row_num, column=5, value=user.gender)
            ws.cell(row=row_num, column=6, value=user.mobile_number)
            ws.cell(row=row_num, column=7, value=user.whatsapp_number)
            ws.cell(row=row_num, column=8, value=user.district)
            ws.cell(row=row_num, column=9, value=user.blood_group)
            ws.cell(row=row_num, column=10, value=user.educational_status)
            ws.cell(row=row_num, column=11, value=user.get_category_display())
            ws.cell(row=row_num, column=12, value=user.university_name)
            ws.cell(row=row_num, column=13, value=user.country_university)
            ws.cell(row=row_num, column=14, value=user.year_of_joining)
            ws.cell(row=row_num, column=15, value=user.year_of_completion)
            ws.cell(row=row_num, column=16, value=user.status)
            ws.cell(
                row=row_num,
                column=17,
                value=user.display_membership_id,
            )

        # Auto-adjust column widths
        from openpyxl.cell.cell import MergedCell
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            
            # If we couldn't find a column letter, skip this column
            if not column_letter:
                continue
            
            # Calculate max length
            for cell in column:
                if not isinstance(cell, MergedCell):
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"members_{category.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Successfully exported {users.count()} members for category: {category}")
        return response
        
    except ImportError:
        logger.error("openpyxl library not installed")
        return JsonResponse({
            'success': False,
            'message': 'Excel export library not installed'
        }, status=500)
    except Exception as e:
        logger.exception(f"Error exporting category members: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required(login_url="/login/")
@admin_required
def download_registrations(request):
    """
    Download registrations based on date filter.

    Args:
        request: HttpRequest object with period and optional date parameters

    Returns:
        HttpResponse: Excel file download

    Logs:
        - Info: Successful download
        - Error: Download failures
    """
    logger.debug("Downloading registrations report")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime, timedelta
        from django.utils import timezone as tz
        
        period = request.GET.get('period', '')
        
        if not period:
            return JsonResponse({
                'success': False,
                'message': 'Period parameter is required'
            }, status=400)
        
        # Calculate date range based on period
        end_date = tz.now()
        
        if period == 'this_month':
            start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            period_label = "This Month"
        elif period == 'last_month':
            first_day_this_month = end_date.replace(day=1)
            last_day_last_month = first_day_this_month - timedelta(days=1)
            start_date = last_day_last_month.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = last_day_last_month.replace(hour=23, minute=59, second=59, microsecond=999999)
            period_label = "Last Month"
        elif period == 'last_3_months':
            start_date = end_date - timedelta(days=90)
            period_label = "Last 3 Months"
        elif period == 'last_6_months':
            start_date = end_date - timedelta(days=180)
            period_label = "Last 6 Months"
        elif period == 'last_year':
            start_date = end_date - timedelta(days=365)
            period_label = "Last 1 Year"
        elif period == 'custom':
            start_date_str = request.GET.get('start_date')
            end_date_str = request.GET.get('end_date')
            
            if not start_date_str or not end_date_str:
                return JsonResponse({
                    'success': False,
                    'message': 'Start date and end date are required for custom period'
                }, status=400)
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = tz.make_aware(start_date.replace(hour=0, minute=0, second=0, microsecond=0))
            
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = tz.make_aware(end_date.replace(hour=23, minute=59, second=59, microsecond=999999))
            
            period_label = f"{start_date_str} to {end_date_str}"
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid period'
            }, status=400)
        
        # Filter users by date_joined
        users = User.objects.filter(
            date_joined__gte=start_date,
            date_joined__lte=end_date
        ).order_by('-date_joined')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        # Add title row
        ws.merge_cells('A1:Q1')
        title_cell = ws['A1']
        title_cell.value = f"Registrations Report - {period_label}"
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Add headers
        headers = [
            'Registration Date', 'Username', 'First Name', 'Last Name', 'Email',
            'Gender', 'Mobile Number', 'WhatsApp Number', 'District', 'Blood Group',
            'Educational Status', 'Category', 'University Name', 'Country',
            'Year of Joining', 'Status',
            'ADAMS membership no.',
        ]
        
        # Style header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for row_num, user in enumerate(users, 3):
            ws.cell(row=row_num, column=1, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=2, value=user.username)
            ws.cell(row=row_num, column=3, value=user.first_name)
            ws.cell(row=row_num, column=4, value=user.last_name)
            ws.cell(row=row_num, column=5, value=user.email)
            ws.cell(row=row_num, column=6, value=user.gender)
            ws.cell(row=row_num, column=7, value=user.mobile_number)
            ws.cell(row=row_num, column=8, value=user.whatsapp_number)
            ws.cell(row=row_num, column=9, value=user.district)
            ws.cell(row=row_num, column=10, value=user.blood_group)
            ws.cell(row=row_num, column=11, value=user.educational_status)
            ws.cell(row=row_num, column=12, value=user.get_category_display())
            ws.cell(row=row_num, column=13, value=user.university_name)
            ws.cell(row=row_num, column=14, value=user.country_university)
            ws.cell(row=row_num, column=15, value=user.year_of_joining)
            ws.cell(row=row_num, column=16, value=user.status)
            ws.cell(
                row=row_num,
                column=17,
                value=user.display_membership_id,
            )

        # Auto-adjust column widths
        from openpyxl.cell.cell import MergedCell
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            
            # If we couldn't find a column letter, skip this column
            if not column_letter:
                continue
            
            # Calculate max length
            for cell in column:
                if not isinstance(cell, MergedCell):
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"registrations_{period}_{tz.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Successfully exported {users.count()} registrations for period: {period_label}")
        return response
        
    except ImportError:
        logger.error("openpyxl library not installed")
        return JsonResponse({
            'success': False,
            'message': 'Excel export library not installed'
        }, status=500)
    except Exception as e:
        logger.exception(f"Error downloading registrations: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@login_required(login_url="/login/")
@admin_required
def send_bulk_email(request):
    """
    Send bulk emails to members of a specific category.
    Uses Server-Sent Events (SSE) to stream progress updates.

    Args:
        request: HttpRequest object with POST data containing category, subject, and message

    Returns:
        StreamingHttpResponse: Progress updates stream

    Logs:
        - Info: Email sending progress
        - Error: Email sending failures
    """
    logger.debug("Starting bulk email send")
    
    def generate():
        try:
            data = json.loads(request.body)
            category = data.get('category')
            subject = data.get('subject')
            message = data.get('message')
            
            if not all([category, subject, message]):
                yield f"data: {json.dumps({'error': 'Missing required fields'})}\n\n"
                return

            allowed_cat = _user_category_values()
            if category != "All" and category not in allowed_cat:
                yield f"data: {json.dumps({'error': 'Invalid category selected.'})}\n\n"
                return

            # Handle "All" categories
            if category == 'All':
                users = User.objects.filter(status='approved')
            else:
                users = User.objects.filter(category=category, status='approved')
            
            total_users = users.count()
            
            if total_users == 0:
                yield f"data: {json.dumps({'complete': True, 'message': 'No approved members found in this category'})}\n\n"
                return
            
            logger.info(f"Sending bulk email to {total_users} members in category: {category}")
            
            sent_count = 0
            failed_count = 0
            
            for index, user in enumerate(users, 1):
                try:
                    # Send individual email
                    send_email_with_attachments(
                        subject=subject,
                        custom_message=message,
                        template_type="custom",
                        recipient_list=[user.email]
                    )
                    
                    sent_count += 1
                    logger.info(f"Email sent to {user.email} ({index}/{total_users})")
                    
                except Exception as e:
                    failed_count += 1
                    logger.error(f"Failed to send email to {user.email}: {str(e)}")
                
                # Calculate progress
                progress = (index / total_users) * 100
                
                # Send progress update
                yield f"data: {json.dumps({'progress': progress, 'message': f'Sent {sent_count} of {total_users} emails...'})}\n\n"
                
                # Small delay to prevent overwhelming the email server
                time.sleep(0.1)
            
            # Send completion message
            completion_message = f"Successfully sent {sent_count} email(s)."
            if failed_count > 0:
                completion_message += f" {failed_count} email(s) failed."
            
            yield f"data: {json.dumps({'complete': True, 'progress': 100, 'message': completion_message})}\n\n"
            logger.info(f"Bulk email completed: {sent_count} sent, {failed_count} failed")
            
        except Exception as e:
            logger.exception(f"Error in bulk email send: {str(e)}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
    
    response = StreamingHttpResponse(generate(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    response['X-Accel-Buffering'] = 'no'
    return response
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
from django.conf import settings
from loguru import logger
import json

from .models import User


@csrf_exempt
@require_POST
def approve_member(request, user_id):
    """
    Approve a member and send welcome email.
    """
    logger.debug(f"Approving member with ID: {user_id}")
    
    try:
        user = User.objects.get(id=user_id)
        
        # Update user status
        user.status = 'approved'
        user.admin_remarks = f"Approved on {user.date_joined.strftime('%Y-%m-%d %H:%M:%S')}"
        user.save()
        
        # Send welcome email
        try:
            subject = "Welcome to ADAMS - Application Approved!"
            message = f"""
Dear {user.first_name} {user.last_name},

Congratulations! Your application to join the Association of Doctors and Medical Students (ADAMS) has been approved!

We are delighted to welcome you to our community. You now have full access to all member benefits and resources.

Login Details:
- Website: {request.build_absolute_uri('/')}
- Username: {user.username}

If you have any questions or need assistance, please don't hesitate to contact us.

Welcome aboard!

Best regards,
ADAMS Team
            """
            
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            logger.info(f"Welcome email sent to {user.email}")
        except Exception as email_error:
            logger.error(f"Failed to send welcome email: {str(email_error)}")
        
        logger.info(f"Member {user.username} approved successfully")
        
        return JsonResponse({
            'success': True,
            'message': 'Member approved successfully'
        })
        
    except User.DoesNotExist:
        logger.error(f"User with ID {user_id} not found")
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except Exception as e:
        logger.exception(f"Error approving member: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def reject_member(request, user_id):
    """
    Reject a member with rejection note and send notification email.
    """
    logger.debug(f"Rejecting member with ID: {user_id}")
    
    try:
        data = json.loads(request.body)
        rejection_note = data.get('rejection_note', '').strip()
        
        if not rejection_note:
            return JsonResponse({
                'success': False,
                'message': 'Rejection note is required'
            }, status=400)
        
        user = User.objects.get(id=user_id)
        
        # Update user status
        user.status = 'rejected'
        user.admin_remarks = f"Rejected: {rejection_note}"
        user.save()
        
        # Send rejection email
        try:
            subject = "ADAMS Membership Application Status"
            message = f"""
Dear {user.first_name} {user.last_name},

Thank you for your interest in joining the Association of Doctors and Medical Students (ADAMS).

After careful review, we regret to inform you that your membership application has not been approved at this time.

Reason: {rejection_note}

If you believe this decision was made in error or if you would like to discuss this further, please feel free to contact us.

We appreciate your interest in ADAMS and wish you all the best in your endeavors.

Best regards,
ADAMS Team
            """
            
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            logger.info(f"Rejection email sent to {user.email}")
        except Exception as email_error:
            logger.error(f"Failed to send rejection email: {str(email_error)}")
        
        logger.info(f"Member {user.username} rejected successfully")
        
        return JsonResponse({
            'success': True,
            'message': 'Member rejected successfully'
        })
        
    except User.DoesNotExist:
        logger.error(f"User with ID {user_id} not found")
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception(f"Error rejecting member: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)



@csrf_exempt
@require_POST
def change_member_status(request, user_id):
    """
    Change member status with admin note.
    """
    logger.debug(f"Changing status for member ID: {user_id}")
    
    try:
        data = json.loads(request.body)
        new_status = data.get('status', '').strip()
        note = data.get('note', '').strip()
        
        if not new_status or new_status not in ['pending', 'approved', 'rejected']:
            return JsonResponse({
                'success': False,
                'message': 'Invalid status'
            }, status=400)
        
        user = User.objects.get(id=user_id)
        old_status = user.status
        
        # Update user status
        user.status = new_status
        if note:
            user.admin_remarks = f"Status changed from {old_status} to {new_status}: {note}"
        else:
            user.admin_remarks = f"Status changed from {old_status} to {new_status}"
        user.save()
        
        logger.info(f"Member {user.username} status changed from {old_status} to {new_status}")
        
        return JsonResponse({
            'success': True,
            'message': 'Status updated successfully'
        })
        
    except User.DoesNotExist:
        logger.error(f"User with ID {user_id} not found")
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception(f"Error changing status: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@login_required(login_url="/login/")
@admin_required
@require_POST
def change_user_role(request, user_id):
    """
    Change user role with admin note.
    """
    logger.debug(f"Changing role for user ID: {user_id}")
    
    try:
        data = json.loads(request.body)
        new_role = data.get('role', '').strip()
        note = data.get('note', '').strip()
        
        # Validate role
        valid_roles = ['admin', 'student', 'intern', 'doctor']
        if not new_role or new_role not in valid_roles:
            return JsonResponse({
                'success': False,
                'message': f'Invalid role. Must be one of: {", ".join(valid_roles)}'
            }, status=400)
        
        user = User.objects.get(id=user_id)
        old_role = user.role
        
        # Prevent changing own role from admin (security measure)
        if user.id == request.user.id and old_role == 'admin' and new_role != 'admin':
            return JsonResponse({
                'success': False,
                'message': 'You cannot change your own role from admin'
            }, status=400)
        
        # Update user role
        user.role = new_role
        if note:
            user.admin_remarks = f"Role changed from {old_role} to {new_role}: {note}"
        else:
            user.admin_remarks = f"Role changed from {old_role} to {new_role}"
        user.save()
        
        logger.info(f"User {user.username} role changed from {old_role} to {new_role} by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'Role updated successfully',
            'new_role': new_role
        })
        
    except User.DoesNotExist:
        logger.error(f"User with ID {user_id} not found")
        return JsonResponse({
            'success': False,
            'message': 'User not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception(f"Error changing role: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def get_notices(request):
    """
    Get all notices for display.
    """
    try:
        from .models import Announcement
        
        # Check if user is authenticated
        is_authenticated = request.user.is_authenticated
        
        # Filter notices based on user authentication
        if is_authenticated:
            # Logged-in users see all notices
            notices = Announcement.objects.all()[:20]
        else:
            # Public users only see public notices
            notices = Announcement.objects.filter(visibility='public')[:20]
        
        notices_data = [{
            'aid': notice.aid,
            'announcement': notice.announcement,
            'hyper_link': notice.hyper_link,
            'visibility': notice.visibility if hasattr(notice, 'visibility') else 'public',
            'date_time': notice.date_time.isoformat() if notice.date_time else notice.date.isoformat()
        } for notice in notices]
        
        return JsonResponse({'success': True, 'notices': notices_data})
    except Exception as e:
        logger.error(f"Error fetching notices: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@csrf_exempt
@require_POST
def add_notice(request):
    """
    Add a new notice to the board.
    """
    try:
        from .models import Announcement
        data = json.loads(request.body)
        announcement_text = data.get('announcement', '').strip()
        hyper_link = data.get('hyper_link', '').strip()
        visibility = data.get('visibility', 'public').strip()
        
        if not announcement_text:
            return JsonResponse({
                'success': False,
                'message': 'Notice message is required'
            }, status=400)
        
        # Get the admin user from request
        user = request.user if request.user.is_authenticated else User.objects.filter(role='admin').first()
        
        notice = Announcement.objects.create(
            uid=user,
            announcement=announcement_text,
            hyper_link=hyper_link if hyper_link else None,
            visibility=visibility
        )
        
        logger.info(f"Notice added by admin: {announcement_text[:50]} (Visibility: {visibility})")
        
        return JsonResponse({
            'success': True,
            'message': 'Notice added successfully',
            'notice_id': notice.aid
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception(f"Error adding notice: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def update_notice(request, notice_id):
    """
    Update an existing notice.
    """
    try:
        from .models import Announcement
        data = json.loads(request.body)
        announcement_text = data.get('announcement', '').strip()
        hyper_link = data.get('hyper_link', '').strip()
        visibility = data.get('visibility', 'public').strip()
        
        if not announcement_text:
            return JsonResponse({
                'success': False,
                'message': 'Notice message is required'
            }, status=400)
        
        notice = Announcement.objects.get(aid=notice_id)
        notice.announcement = announcement_text
        notice.hyper_link = hyper_link if hyper_link else None
        notice.visibility = visibility
        notice.save()
        
        logger.info(f"Notice {notice_id} updated (Visibility: {visibility})")
        
        return JsonResponse({
            'success': True,
            'message': 'Notice updated successfully'
        })
        
    except Announcement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Notice not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception(f"Error updating notice: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def delete_notice(request, notice_id):
    """
    Delete a notice from the board.
    """
    try:
        from .models import Announcement
        notice = Announcement.objects.get(aid=notice_id)
        notice.delete()
        
        logger.info(f"Notice {notice_id} deleted")
        
        return JsonResponse({
            'success': True,
            'message': 'Notice deleted successfully'
        })
        
    except Announcement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Notice not found'
        }, status=404)
    except Exception as e:
        logger.exception(f"Error deleting notice: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)



@login_required(login_url="/login/")
def notice_board_view(request):
    """
    Render the notice board page for logged-in users.
    
    This view displays all announcements/notices to logged-in users.
    """
    logger.debug("Processing notice board page request")
    
    uname = request.session.get("username")
    user = None
    
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the notice board")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")
    
    data = {"user": user}
    logger.debug("Rendering notice board page")
    return render(request, "notice_board.html", {"data": data})



def get_messages(request):
    """Get all contact messages with optional filtering."""
    try:
        from .models import ContactMessage
        filter_type = request.GET.get('filter', 'all')
        
        if filter_type == 'all':
            messages = ContactMessage.objects.all()[:50]
        else:
            messages = ContactMessage.objects.filter(status=filter_type)[:50]
        
        messages_data = [{
            'id': msg.id,
            'name': msg.name,
            'email': msg.email,
            'subject': msg.subject,
            'message': msg.message,
            'status': msg.status,
            'created_at': msg.created_at.isoformat()
        } for msg in messages]
        
        return JsonResponse({'success': True, 'messages': messages_data})
    except Exception as e:
        logger.error(f"Error fetching messages: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def get_message(request, message_id):
    """Get a single contact message by ID."""
    try:
        from .models import ContactMessage
        message = ContactMessage.objects.get(id=message_id)
        
        message_data = {
            'id': message.id,
            'name': message.name,
            'email': message.email,
            'subject': message.subject,
            'message': message.message,
            'status': message.status,
            'reply': message.reply,
            'created_at': message.created_at.isoformat(),
            'replied_at': message.replied_at.isoformat() if message.replied_at else None
        }
        
        return JsonResponse({'success': True, 'message': message_data})
    except ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Message not found'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching message: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def mark_message_read(request, message_id):
    """Mark a contact message as read."""
    try:
        from .models import ContactMessage
        message = ContactMessage.objects.get(id=message_id)
        if message.status == 'new':
            message.status = 'read'
            message.save()
        
        return JsonResponse({'success': True})
    except ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Message not found'}, status=404)
    except Exception as e:
        logger.error(f"Error marking message as read: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def reply_message(request, message_id):
    """Send a reply to a contact message via email."""
    try:
        from .models import ContactMessage
        from django.utils import timezone
        
        data = json.loads(request.body)
        reply_text = data.get('reply', '').strip()
        
        if not reply_text:
            return JsonResponse({'success': False, 'message': 'Reply text is required'}, status=400)
        
        message = ContactMessage.objects.get(id=message_id)
        
        # Save reply to database
        message.reply = reply_text
        message.status = 'replied'
        message.replied_at = timezone.now()
        message.replied_by = request.user if request.user.is_authenticated else None
        message.save()
        
        # Send email reply
        email_subject = f"Re: {message.subject}"
        email_body = f"""
Dear {message.name},

Thank you for contacting us. Here is our response to your message:

{reply_text}

---
Your Original Message:
{message.message}

---
Best regards,
Association Of Doctors And Medical Students (ADAMS)
"""
        
        try:
            send_mail(
                subject=email_subject,
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[message.email],
                fail_silently=False
            )
            logger.info(f"Reply sent to {message.email} for message ID {message_id}")
        except Exception as email_error:
            logger.error(f"Error sending reply email: {str(email_error)}")
            return JsonResponse({
                'success': False,
                'message': f'Reply saved but email failed to send: {str(email_error)}'
            }, status=500)
        
        return JsonResponse({'success': True, 'message': 'Reply sent successfully'})
        
    except ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Message not found'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.exception(f"Error replying to message: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def delete_message(request, message_id):
    """Delete a contact message."""
    try:
        from .models import ContactMessage
        message = ContactMessage.objects.get(id=message_id)
        message.delete()
        
        logger.info(f"Message {message_id} deleted")
        return JsonResponse({'success': True, 'message': 'Message deleted successfully'})
        
    except ContactMessage.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Message not found'}, status=404)
    except Exception as e:
        logger.exception(f"Error deleting message: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



@login_required(login_url="/login/")
def rulebook_viewer(request):
    """Render the rulebook viewer page for logged-in users."""
    logger.debug("Processing rulebook viewer request")
    
    uname = request.session.get("username")
    user = None
    
    if uname:
        try:
            user = User.objects.get(username=uname)
            logger.info(f"User {uname} accessed the rulebook viewer")
        except User.DoesNotExist:
            logger.warning(f"User {uname} from session not found in database")
            user = None
    else:
        logger.debug("No username found in session")
    
    data = {"user": user}
    logger.debug("Rendering rulebook viewer page")
    return render(request, "rulebook_viewer.html", {"data": data})


def get_rulebooks(request):
    """Get all rulebooks for admin management."""
    try:
        from .models import Rulebook
        rulebooks = Rulebook.objects.all()[:10]
        
        rulebooks_data = [{
            'id': rb.id,
            'title': rb.title,
            'description': rb.description,
            'pdf_file': rb.pdf_file,
            'is_active': rb.is_active,
            'uploaded_at': rb.uploaded_at.isoformat(),
            'uploaded_by': rb.uploaded_by.username if rb.uploaded_by else 'Admin'
        } for rb in rulebooks]
        
        return JsonResponse({'success': True, 'rulebooks': rulebooks_data})
    except Exception as e:
        logger.error(f"Error fetching rulebooks: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def get_active_rulebook(request):
    """Get the active rulebook for viewing."""
    try:
        from .models import Rulebook
        rulebook = Rulebook.objects.filter(is_active=True).first()
        
        if rulebook:
            try:
                # Always use Django proxy so PDF.js works for both local files and MinIO objects.
                pdf_url = request.build_absolute_uri(reverse("serve_rulebook_pdf"))

                return JsonResponse({
                    'success': True,
                    'rulebook': {
                        'id': rulebook.id,
                        'title': rulebook.title,
                        'description': rulebook.description,
                        'pdf_url': pdf_url
                    }
                })
            except Exception as e:
                logger.error(f"Error generating presigned URL: {str(e)}")
                return JsonResponse({'success': False, 'message': 'Error accessing PDF file'}, status=500)
        else:
            return JsonResponse({'success': False, 'message': 'No active rulebook found'}, status=404)
    except Exception as e:
        logger.error(f"Error fetching active rulebook: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def upload_rulebook(request):
    """Upload a new rulebook PDF."""
    import uuid

    try:
        from .models import Rulebook

        title = request.POST.get("title", "ADAMS Rulebook").strip() or "ADAMS Rulebook"
        description = (request.POST.get("description") or "").strip()
        raw_active = (request.POST.get("is_active") or "").strip().lower()
        is_active = raw_active in ("true", "1", "on", "yes")
        pdf_file = request.FILES.get("pdf_file")

        if not pdf_file:
            return JsonResponse({"success": False, "message": "No PDF file provided"}, status=400)

        if not (pdf_file.name or "").lower().endswith(".pdf"):
            return JsonResponse({"success": False, "message": "File must be a PDF"}, status=400)

        safe_pdf_name = os.path.basename(pdf_file.name or "rulebook.pdf").replace("..", "_")
        object_key = f"rulebooks/{uuid.uuid4()}_{safe_pdf_name}"

        if getattr(settings, "MINIO_ENABLED", False):
            pdf_file.seek(0)
            if not upload_file_to_minio(pdf_file, object_key, "application/pdf"):
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Could not upload to MinIO. Check MINIO_* settings and server logs.",
                    },
                    status=500,
                )
            db_file_name = f"minio:{object_key}"
            logger.info("Uploaded rulebook PDF to MinIO: %s", object_key)
        else:
            media_path = os.path.join(settings.MEDIA_ROOT, object_key)
            os.makedirs(os.path.dirname(media_path), exist_ok=True)
            try:
                pdf_file.seek(0)
                with open(media_path, "wb+") as destination:
                    for chunk in pdf_file.chunks():
                        destination.write(chunk)
                db_file_name = f"media/{object_key}"
                logger.info("Uploaded rulebook PDF to local storage: %s", object_key)
            except Exception as e:
                logger.error("Error uploading rulebook to local storage: %s", e)
                return JsonResponse(
                    {"success": False, "message": "Error uploading file"},
                    status=500,
                )
        
        # If setting as active, deactivate all other rulebooks
        if is_active:
            Rulebook.objects.all().update(is_active=False)
        
        # Create rulebook record
        user = request.user if request.user.is_authenticated else None
        rulebook = Rulebook.objects.create(
            title=title,
            description=description,
            pdf_file=db_file_name,  # Store relative path with media/ prefix
            is_active=is_active,
            uploaded_by=user
        )
        
        logger.info(f"Created rulebook: {title}")
        
        return JsonResponse({
            'success': True,
            'message': 'Rulebook uploaded successfully',
            'rulebook_id': rulebook.id
        })
        
    except Exception as e:
        logger.exception(f"Error uploading rulebook: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def toggle_rulebook(request, rulebook_id):
    """Toggle rulebook active status."""
    try:
        from .models import Rulebook
        data = json.loads(request.body)
        is_active = data.get('is_active', False)
        
        rulebook = Rulebook.objects.get(id=rulebook_id)
        
        # If activating, deactivate all others
        if is_active:
            Rulebook.objects.all().update(is_active=False)
        
        rulebook.is_active = is_active
        rulebook.save()
        
        logger.info(f"Rulebook {rulebook_id} {'activated' if is_active else 'deactivated'}")
        
        return JsonResponse({'success': True})
        
    except Rulebook.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Rulebook not found'}, status=404)
    except Exception as e:
        logger.exception(f"Error toggling rulebook: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@require_POST
def delete_rulebook(request, rulebook_id):
    """Delete a rulebook."""
    try:
        from .models import Rulebook
        
        rulebook = Rulebook.objects.get(id=rulebook_id)

        try:
            stored = str(rulebook.pdf_file or "")
            if stored.startswith("minio:"):
                object_name = stored[len("minio:") :].lstrip("/")
                if object_name:
                    delete_file_from_minio(object_name)
                    logger.info("Deleted rulebook PDF from MinIO: %s", object_name)
            else:
                pdf_path = stored
                if not pdf_path.startswith("media/"):
                    pdf_path = f"media/{pdf_path}"
                full_path = os.path.join(settings.BASE_DIR, pdf_path)
                full_path = os.path.normpath(full_path)
                media_root = os.path.normpath(settings.MEDIA_ROOT)
                if full_path.startswith(media_root) and os.path.exists(full_path):
                    os.remove(full_path)
                    logger.info("Deleted rulebook PDF from local storage: %s", rulebook.pdf_file)
        except Exception as e:
            logger.warning("Error deleting rulebook file: %s", e)
        
        # Delete from database
        rulebook.delete()
        logger.info(f"Deleted rulebook {rulebook_id}")
        
        return JsonResponse({'success': True, 'message': 'Rulebook deleted successfully'})
        
    except Rulebook.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Rulebook not found'}, status=404)
    except Exception as e:
        logger.exception(f"Error deleting rulebook: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)



def serve_rulebook_pdf(request):
    """Serve the active rulebook PDF through Django as a proxy."""
    try:
        from .models import Rulebook
        from django.http import HttpResponse

        rulebook = Rulebook.objects.filter(is_active=True).first()
        
        if not rulebook:
            return HttpResponse('No active rulebook found', status=404)
        
        try:
            stored = str(rulebook.pdf_file or "")
            if stored.startswith("minio:"):
                object_name = stored[len("minio:") :].lstrip("/")
                if not object_name:
                    return HttpResponse("PDF reference invalid", status=404)
                pdf_data = fetch_file_from_minio(object_name)
                if not pdf_data:
                    return HttpResponse("PDF file not found", status=404)
            else:
                pdf_path = stored
                if not pdf_path.startswith("media/"):
                    pdf_path = f"media/{pdf_path}"
                full_path = os.path.join(settings.BASE_DIR, pdf_path)
                full_path = os.path.normpath(full_path)
                media_root = os.path.normpath(settings.MEDIA_ROOT)
                if not full_path.startswith(media_root):
                    return HttpResponse("Access denied", status=403)
                if not os.path.exists(full_path):
                    return HttpResponse("PDF file not found", status=404)
                with open(full_path, "rb") as f:
                    pdf_data = f.read()

            http_response = HttpResponse(pdf_data, content_type="application/pdf")
            http_response['Content-Disposition'] = f'inline; filename="{rulebook.title}.pdf"'
            http_response['Access-Control-Allow-Origin'] = '*'
            
            return http_response
            
        except Exception as e:
            logger.error(f"Error serving PDF: {str(e)}")
            return HttpResponse('Error loading PDF', status=500)
            
    except Exception as e:
        logger.exception(f"Error in serve_rulebook_pdf: {str(e)}")
        return HttpResponse('Server error', status=500)



def privacy_policy(request):
    """Render the privacy policy page."""
    uname = request.session.get("username")
    user = None
    if uname:
        try:
            user = User.objects.get(username=uname)
        except User.DoesNotExist:
            user = None
    
    data = {"user": user}
    return render(request, "privacy_policy.html", {"data": data})


def terms_conditions(request):
    """Render the terms and conditions page."""
    uname = request.session.get("username")
    user = None
    if uname:
        try:
            user = User.objects.get(username=uname)
        except User.DoesNotExist:
            user = None
    
    data = {"user": user}
    return render(request, "terms_conditions.html", {"data": data})


def refund_policy(request):
    """Render the refund and cancellation policy page."""
    uname = request.session.get("username")
    user = None
    if uname:
        try:
            user = User.objects.get(username=uname)
        except User.DoesNotExist:
            user = None
    
    data = {"user": user}
    return render(request, "refund_policy.html", {"data": data})


# Payment Gateway Integration - UatPayments
def payment_request(request):
    """
    Handle payment request and redirect to payment gateway.
    """
    # UatPayments API credentials
    API_KEY = "fb6bca86-b429-4abf-a42f-824bdd29022e"
    SALT = "80c67bfdf027da08de88ab5ba903fecafaab8f6d"
    URL = "https://pgbiz.omniware.in/v2/paymentrequest"
    
    posted = {}
    for i in request.POST:
        posted[i] = request.POST[i]
    
    # Hash sequence for payment request
    hash_sequence = "address_line_1|address_line_2|amount|api_key|city|country|currency|description|email|mode|name|order_id|phone|return_url|state|udf1|udf2|udf3|udf4|udf5|zip_code"
    
    # Add API_KEY to posted data for hash generation
    posted['api_key'] = API_KEY
    
    hash_string = SALT
    hash_vars_seq = hash_sequence.split('|')
    
    for i in hash_vars_seq:
        if i in posted.keys():
            if len(str(posted[i])) > 0:
                hash_string += '|'
                hash_string += str(posted[i])
    
    hash_value = hashlib.sha512(hash_string.encode()).hexdigest().upper()
    action = URL
    
    # Validate required fields
    required_fields = [
        'amount', 'address_line_1', 'city', 'name', 'email', 
        'phone', 'order_id', 'currency', 'description', 
        'country', 'return_url'
    ]
    
    if all(posted.get(field) and posted.get(field) != '' for field in required_fields):
        return render(request, 'payment_request.html', {
            "posted": posted,
            "hash": hash_value,
            "API_KEY": API_KEY,
            "action": action
        })
    else:
        hash_value = ''
        return render(request, 'payment_request.html', {
            "posted": posted,
            "hash": hash_value,
            "API_KEY": API_KEY,
            "action": "."
        })


@csrf_exempt
def payment_response(request):
    """
    Handle payment response from payment gateway.
    """
    if request.method != 'POST':
        return render(request, 'payment_failure.html', {
            "response_message": "Invalid request method"
        })
    
    try:
        response_code = request.POST.get("response_code", "")
        transaction_id = request.POST.get("transaction_id", "")
        response_message = request.POST.get("response_message", "")
        amount = request.POST.get("amount", "")
        hash_value = request.POST.get("hash", "")
        
        # UatPayments SALT for response verification
        SALT = "80c67bfdf027da08de88ab5ba903fecafaab8f6d"
        hash_string = SALT
        
        # Build hash string from POST data (excluding hash itself)
        for i in sorted(request.POST):
            if i != 'hash':
                if len(request.POST[i]) > 0:
                    hash_string += '|'
                    hash_string += request.POST[i]
        
        calculated_hash = hashlib.sha512(hash_string.encode()).hexdigest().upper()
        
        # Verify hash
        if hash_value == calculated_hash:
            if response_code == '0':
                # Payment successful
                return render(request, 'payment_success.html', {
                    "txnid": transaction_id,
                    "status": response_message,
                    "amount": amount
                })
            else:
                # Payment failed
                return render(request, 'payment_failure.html', {
                    "response_message": response_message or "Transaction Failed"
                })
        else:
            # Hash mismatch
            return render(request, 'payment_failure.html', {
                "response_message": "Hash Mismatched, Transaction Failed"
            })
    except Exception as e:
        logger.error(f"Error processing payment response: {e}")
        return render(request, 'payment_failure.html', {
            "response_message": "Error processing payment response"
        })


# ============================================
# Registration Questions Management Views
# ============================================

@login_required(login_url="/login/")
@admin_required
def manage_registration_questions(request):
    """Admin view to manage registration questions"""
    questions = RegistrationQuestion.objects.all().order_by('order', 'created_at')
    return render(request, 'manage_registration_questions.html', {
        'questions': questions
    })


@login_required(login_url="/login/")
@admin_required
@require_POST
def add_registration_question(request):
    """Add a new registration question"""
    try:
        question_text = request.POST.get('question_text')
        question_type = request.POST.get('question_type', 'text')
        is_required = request.POST.get('is_required') == 'on'
        order = int(request.POST.get('order', 0))
        options = request.POST.get('options', '').strip()
        help_text = request.POST.get('help_text', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        question = RegistrationQuestion.objects.create(
            question_text=question_text,
            question_type=question_type,
            is_required=is_required,
            order=order,
            options=options,
            help_text=help_text,
            is_active=is_active,
            created_by=request.user
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Question added successfully',
            'question_id': question.id
        })
    except Exception as e:
        logger.error(f"Error adding registration question: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required(login_url="/login/")
@admin_required
@require_POST
def update_registration_question(request, question_id):
    """Update an existing registration question"""
    try:
        question = get_object_or_404(RegistrationQuestion, id=question_id)
        
        question.question_text = request.POST.get('question_text')
        question.question_type = request.POST.get('question_type', 'text')
        question.is_required = request.POST.get('is_required') == 'on'
        question.order = int(request.POST.get('order', 0))
        question.options = request.POST.get('options', '').strip()
        question.help_text = request.POST.get('help_text', '').strip()
        question.is_active = request.POST.get('is_active') == 'on'
        question.save()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Question updated successfully'
        })
    except Exception as e:
        logger.error(f"Error updating registration question: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required(login_url="/login/")
@admin_required
@require_POST
def delete_registration_question(request, question_id):
    """Delete a registration question"""
    try:
        question = get_object_or_404(RegistrationQuestion, id=question_id)
        question.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': 'Question deleted successfully'
        })
    except Exception as e:
        logger.error(f"Error deleting registration question: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required(login_url="/login/")
@admin_required
def get_registration_question(request, question_id):
    """Get a single registration question (for editing)"""
    try:
        question = get_object_or_404(RegistrationQuestion, id=question_id)
        return JsonResponse({
            'status': 'success',
            'question': {
                'id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'is_required': question.is_required,
                'order': question.order,
                'options': question.options or '',
                'help_text': question.help_text or '',
                'is_active': question.is_active,
            }
        })
    except Exception as e:
        logger.error(f"Error getting registration question: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


@login_required(login_url="/login/")
@admin_required
@require_POST
def migrate_existing_fields_to_questions(request):
    """Migrate existing registration form fields to custom questions"""
    try:
        # Define field mappings: (field_name, question_text, question_type, is_required, order, help_text, options)
        field_mappings = [
            ('username', 'Username', 'text', True, 1, 'Choose a unique username', None),
            ('first_name', 'First Name', 'text', True, 2, None, None),
            ('middle_name', 'Middle Name', 'text', False, 3, None, None),
            ('last_name', 'Last Name', 'text', True, 4, None, None),
            ('gender', 'Gender', 'dropdown', True, 5, None, 'Male,Female,Other'),
            ('whatsapp_number', 'WhatsApp Number', 'tel', True, 6, 'Include country code (e.g., +91)', None),
            ('mobile_number', 'Mobile Number', 'tel', True, 7, 'Include country code (e.g., +91)', None),
            ('address_communication', 'Address for Communication', 'textarea', True, 8, None, None),
            ('address_permanent', 'Permanent Address', 'textarea', True, 9, None, None),
            ('district', 'District', 'text', True, 10, None, None),
            ('father_spouse_details', 'Father/Spouse Name & Occupation', 'textarea', True, 11, None, None),
            ('blood_group', 'Blood Group', 'dropdown', True, 12, None, 'A+,A-,B+,B-,AB+,AB-,O+,O-,Bombay Group'),
            ('educational_status', 'Educational Status', 'dropdown', True, 13, None, 'Student (@Abroad University),Graduated / FMGE Aspirant,FMGE Passed Candidate,Internship,Working Doctor,Post Graduation ongoing,Post Graduation Completed, Working'),
            ('category', 'Category You Belong To For Membership', 'dropdown', True, 14, None, 'Student,Graduate/Trying FMGE,FMGE passed,Working Doctor (With SMC/NMC Registration),Working PG Doctor (With SMC/NMC Registration)'),
            ('university_name', 'University Name', 'textarea', True, 15, 'Full name of university (short forms not accepted)', None),
            ('country_university', 'Country Where University is situated', 'textarea', True, 16, None, None),
            ('year_of_joining', 'Year of Joining Medical Education', 'date', True, 17, None, None),
            ('year_of_completion', 'Year of Completion/Expected Completion of course', 'date', True, 18, None, None),
            ('photo', 'Passport Size Photo', 'file', True, 19, 'Upload front facing photograph. It will be used for Association ID card', None),
            ('state_nmc', 'State/NMC registration certificate', 'file', False, 20, 'For registered doctors only', None),
            ('passport', 'Passport (Front and Back)', 'file', True, 21, None, None),
            ('medical_qualification', 'Medical qualification (Provisional/Permanent) Degree/ Diploma Certificate/ Student ID Card', 'file', True, 22, None, None),
            ('date_time_of_payment', 'Date and time of Payment (Transaction)', 'date', True, 23, None, None),
            ('payment_transaction_proof', 'Payment Transaction Proof/screenshot', 'file', True, 24, 'Upload screenshot or slip containing transaction number', None),
            ('willing_to_be_donor', 'Willing to be a blood donor?', 'checkbox', False, 25, None, 'Yes'),
            ('mid', 'ADAMS membership number', 'text', False, 26, 'Bulk XLSX: value in `mid` column is stored after import. Web signup: left blank, auto ADAMS-{pk} on save.', None),
            ('agreement', 'Agreement for Joining Association of Doctors And Medical Students (ADAMS)', 'checkbox', True, 27, None, 'I agree'),
            ('application', 'I agree to the terms and conditions of the association and I hereby submit my application for membership in ADAMS', 'checkbox', True, 28, None, 'I agree'),
        ]
        
        created_count = 0
        skipped_count = 0
        
        for field_name, question_text, question_type, is_required, order, help_text, options in field_mappings:
            # Skip username, email, password fields as per user requirement
            if field_name in ['username']:
                skipped_count += 1
                continue
            
            # Check if question already exists
            if RegistrationQuestion.objects.filter(question_text=question_text).exists():
                skipped_count += 1
                continue
            
            RegistrationQuestion.objects.create(
                question_text=question_text,
                question_type=question_type,
                is_required=is_required,
                order=order,
                options=options,
                help_text=help_text,
                is_active=True,
                created_by=request.user
            )
            created_count += 1
        
        return JsonResponse({
            'status': 'success',
            'message': f'Migration completed. Created {created_count} questions, skipped {skipped_count} existing questions.',
            'created': created_count,
            'skipped': skipped_count
        })
    except Exception as e:
        logger.error(f"Error migrating fields to questions: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)


# ============================================
# System Settings Management Views
# ============================================

@login_required(login_url="/login/")
@admin_required
def manage_system_settings(request):
    """Admin view to manage system settings"""
    # Initialize default 2FA setting if it doesn't exist
    if not SystemSettings.objects.filter(key='two_factor_enabled').exists():
        SystemSettings.set_setting(
            key='two_factor_enabled',
            value='true',
            description='Enable or disable two-factor authentication (OTP) for user login'
        )
    
    two_factor_enabled = SystemSettings.get_setting('two_factor_enabled', default='true')
    two_factor_enabled = str(two_factor_enabled).lower() in ['true', '1', 'yes', 'on']
    
    return render(request, 'manage_system_settings.html', {
        'two_factor_enabled': two_factor_enabled
    })


@login_required(login_url="/login/")
@admin_required
@require_POST
def toggle_two_factor_auth(request):
    """Toggle 2FA on/off"""
    try:
        enabled = request.POST.get('enabled', 'false').lower() in ['true', '1', 'yes', 'on']
        
        SystemSettings.set_setting(
            key='two_factor_enabled',
            value='true' if enabled else 'false',
            description='Enable or disable two-factor authentication (OTP) for user login',
            user=request.user
        )
        
        status_text = 'enabled' if enabled else 'disabled'
        logger.info(f"2FA {status_text} by admin {request.user.username}")
        
        return JsonResponse({
            'status': 'success',
            'message': f'Two-factor authentication has been {status_text}.',
            'enabled': enabled
        })
    except Exception as e:
        logger.error(f"Error toggling 2FA: {e}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
