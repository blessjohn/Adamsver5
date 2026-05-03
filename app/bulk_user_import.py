"""
Admin bulk user import from XLSX (matches registration form core fields).
Document file paths are set to a placeholder until members upload real files via profile.
"""

from datetime import date, datetime
from io import BytesIO

import openpyxl
from django.db import IntegrityError, transaction
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill

from .models import User
from .utils import generate_random_password

# Paths stored on User for uploads not supplied in bulk import (replace via member profile later).
DOC_PLACEHOLDER = "bulk-import/pending"

BULK_USER_HEADERS = (
    "username",
    "password",
    "email",
    "first_name",
    "middle_name",
    "last_name",
    "gender",
    "whatsapp_number",
    "mobile_number",
    "address_communication",
    "address_permanent",
    "district",
    "father_spouse_details",
    "blood_group",
    "role",
    "educational_status",
    "category",
    "university_name",
    "country_university",
    "year_of_joining",
    "year_of_completion",
    "date_time_of_payment",
    "status",
    "willing_to_be_donor",
    "mid",
    "admin_remarks",
)


def _cell_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, bool):
        return "true" if val else "false"
    return str(val).strip()


def _truthy(val: str) -> bool:
    return val.lower() in ("1", "true", "yes", "y", "on")


def _allowed_sets():
    return {
        "gender": {c[0] for c in User.gender_choices},
        "blood_group": {c[0] for c in User.blood_group_choices},
        "role": {c[0] for c in User._meta.get_field("role").choices},
        "educational_status": {c[0] for c in User._meta.get_field("educational_status").choices},
        "category": {c[0] for c in User._meta.get_field("category").choices},
        "status": {c[0] for c in User._meta.get_field("status").choices},
    }


def build_bulk_user_template_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, key in enumerate(BULK_USER_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=key)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    note = ws.cell(row=2, column=1, value="Add one user per row (do not change row 1 headers).")
    note.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(BULK_USER_HEADERS))

    ws.freeze_panes = "A3"

    ref = wb.create_sheet("Field_reference", 1)
    ref.append(["Column (row 1 on Users)", "Required", "Notes"])
    ref_header = ref[1]
    for cell in ref_header:
        cell.fill = header_fill
        cell.font = header_font

    notes = [
        ("username", "Yes", "Unique login; letters, digits, @ . + - _"),
        ("password", "No", "Leave empty to auto-generate a random password."),
        ("email", "Yes", "Unique."),
        ("first_name / middle_name / last_name", "No", ""),
        ("gender", "Yes", f"One of: {', '.join(sorted(_allowed_sets()['gender']))}"),
        ("whatsapp_number", "Yes", "Unique; digits."),
        ("mobile_number", "Yes", "Unique; digits."),
        ("address_communication / address_permanent", "Yes", ""),
        ("district", "Yes", "District name or custom text if Others."),
        ("father_spouse_details", "Yes", ""),
        ("blood_group", "Yes", f"One of: {', '.join(sorted(_allowed_sets()['blood_group']))}"),
        ("role", "No", f"Default student. One of: {', '.join(sorted(_allowed_sets()['role']))}"),
        ("educational_status", "Yes", "Exact value from registration form list."),
        ("category", "Yes", "Exact stored value (see User model / registration)."),
        ("university_name / country_university", "Yes", ""),
        ("year_of_joining / year_of_completion", "Yes", "Date or text as on form."),
        ("date_time_of_payment", "Yes", "Text describing payment time or bulk note."),
        ("status", "No", f"Default pending. One of: {', '.join(sorted(_allowed_sets()['status']))}"),
        ("willing_to_be_donor", "No", "true/false/yes/no/1/0"),
        ("mid", "No", "ADAMS membership number from the sheet is saved to the user after import. Leave blank to use auto ADAMS-{pk}."),
        ("admin_remarks", "No", "Internal note."),
        ("", "", ""),
        ("Documents", "N/A", f"Photo, passport, medical qualification, payment proof are set to '{DOC_PLACEHOLDER}'. Members should upload real files from their profile when possible."),
    ]
    for row in notes:
        ref.append(row)
    ref.column_dimensions["A"].width = 28
    ref.column_dimensions["B"].width = 12
    ref.column_dimensions["C"].width = 72

    allowed = wb.create_sheet("Allowed_pick_values", 2)
    allowed.append(["Field", "Allowed value (exact)"])
    ah = allowed[1]
    for cell in ah:
        cell.fill = header_fill
        cell.font = header_font
    for field in ("gender", "blood_group", "role", "educational_status", "category", "status"):
        for val in sorted(_allowed_sets()[field]):
            allowed.append([field, val])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def import_users_from_xlsx(file_obj):
    """
    Returns (created_count, error_messages, created_accounts).
    ``created_accounts`` is a list of dicts with username, email, plain_password for welcome email.
    Each successful user is saved in its own transaction.
    """
    allowed = _allowed_sets()
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if "Users" in wb.sheetnames:
        ws = wb["Users"]
    else:
        ws = wb.active

    rows = ws.iter_rows(min_row=1, values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return 0, ["The spreadsheet is empty."], []

    headers = [_cell_str(h).lower() for h in header_row]
    if headers != list(BULK_USER_HEADERS):
        return 0, [
            "Header row must match the template exactly (column order and names). "
            "Download a fresh template and paste your rows under row 1 without changing headers."
        ], []

    created = 0
    errors: list[str] = []
    created_accounts: list[dict] = []
    row_index = 1

    for row in rows:
        row_index += 1
        if row is None or all(v is None or _cell_str(v) == "" for v in row):
            continue

        cells = list(row) + [None] * (len(BULK_USER_HEADERS) - len(row))
        data = {BULK_USER_HEADERS[i]: _cell_str(cells[i]) for i in range(len(BULK_USER_HEADERS))}

        # Skip blank / template hint rows (e.g. merged instruction row under headers).
        if not data["email"] and not data["mobile_number"] and not data["whatsapp_number"]:
            continue

        username = data["username"]
        email = data["email"].lower()
        if not username or not email:
            errors.append(f"Row {row_index}: username and email are required.")
            continue

        pwd = data["password"] or generate_random_password(12)

        gender = data["gender"]
        if gender not in allowed["gender"]:
            errors.append(f"Row {row_index}: invalid gender '{gender}'.")
            continue

        blood = data["blood_group"]
        if blood not in allowed["blood_group"]:
            errors.append(f"Row {row_index}: invalid blood_group '{blood}'.")
            continue

        edu = data["educational_status"]
        if edu not in allowed["educational_status"]:
            errors.append(f"Row {row_index}: invalid educational_status '{edu}'.")
            continue

        cat = data["category"]
        if cat not in allowed["category"]:
            errors.append(f"Row {row_index}: invalid category '{cat}'.")
            continue

        role = data["role"] or "student"
        if role not in allowed["role"]:
            errors.append(f"Row {row_index}: invalid role '{role}'.")
            continue

        status = data["status"] or "pending"
        if status not in allowed["status"]:
            errors.append(f"Row {row_index}: invalid status '{status}'.")
            continue

        required = [
            "whatsapp_number",
            "mobile_number",
            "address_communication",
            "address_permanent",
            "district",
            "father_spouse_details",
            "university_name",
            "country_university",
            "year_of_joining",
            "year_of_completion",
            "date_time_of_payment",
        ]
        missing = [k for k in required if not data[k]]
        if missing:
            errors.append(f"Row {row_index}: missing required fields: {', '.join(missing)}.")
            continue

        if User.objects.filter(username__iexact=username).exists():
            errors.append(f"Row {row_index}: username '{username}' already exists.")
            continue
        if User.objects.filter(email__iexact=email).exists():
            errors.append(f"Row {row_index}: email '{email}' already exists.")
            continue
        if User.objects.filter(mobile_number=data["mobile_number"]).exists():
            errors.append(f"Row {row_index}: mobile_number already registered.")
            continue
        if User.objects.filter(whatsapp_number=data["whatsapp_number"]).exists():
            errors.append(f"Row {row_index}: whatsapp_number already registered.")
            continue

        willing = _truthy(data["willing_to_be_donor"]) if data["willing_to_be_donor"] else False

        mid_from_sheet = (data["mid"] or "").strip()[:100]
        mid_initial = mid_from_sheet or None

        try:
            with transaction.atomic():
                user = User(
                    username=username[:150],
                    email=email[:254],
                    first_name=(data["first_name"] or "")[:150],
                    middle_name=(
                        (data["middle_name"] or "")[:150] or None
                    ),
                    last_name=(data["last_name"] or "")[:150],
                    gender=gender,
                    whatsapp_number=data["whatsapp_number"][:15],
                    mobile_number=data["mobile_number"][:15],
                    address_communication=data["address_communication"],
                    address_permanent=data["address_permanent"],
                    district=data["district"][:50],
                    father_spouse_details=data["father_spouse_details"],
                    blood_group=blood,
                    role=role,
                    educational_status=edu,
                    category=cat,
                    university_name=data["university_name"][:100],
                    country_university=data["country_university"][:100],
                    year_of_joining=data["year_of_joining"][:100],
                    year_of_completion=data["year_of_completion"][:100],
                    photo=DOC_PLACEHOLDER,
                    state_nmc="",
                    passport=DOC_PLACEHOLDER,
                    medical_qualification=DOC_PLACEHOLDER,
                    date_time_of_payment=data["date_time_of_payment"][:100],
                    payment_transaction_proof=DOC_PLACEHOLDER,
                    willing_to_be_donor=willing,
                    agreement=True,
                    application=True,
                    mid=mid_initial,
                    status=status,
                    admin_remarks=data["admin_remarks"] or "",
                    is_active=True,
                )
                user.set_password(pwd)
                user.save()
            created += 1
            created_accounts.append(
                {"username": user.username, "email": user.email, "plain_password": pwd}
            )
            logger.info(f"Bulk import created user id={user.id} username={username}")
        except IntegrityError as e:
            errors.append(f"Row {row_index}: database constraint — {e}")
        except Exception as e:
            logger.exception(f"Bulk import row {row_index} failed")
            errors.append(f"Row {row_index}: {e}")

    return created, errors, created_accounts
